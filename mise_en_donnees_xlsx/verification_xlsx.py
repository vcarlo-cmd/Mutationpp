#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Vérification du classeur `mise_en_donnees_materiaux.xlsx`.

Trois contrôles, tous rejouables :

  1. TOUTES les formules du classeur sont évaluées (moteur `formulas`, pur
     Python) : aucune ne doit renvoyer d'erreur Excel (#REF!, #NAME?, #DIV/0!…).
  2. La composition calculée par la chaîne de chaque onglet est comparée à la
     composition RÉELLEMENT lue dans `data/mixtures/*.xml`.
  3. Les valeurs calculées sont réinjectées dans le fichier (cache <v>), afin
     que le classeur s'ouvre déjà renseigné et soit lisible par pandas ou par
     un aperçu, sans passer par un recalcul.

Usage :
    python verification_xlsx.py

Prérequis : pip install openpyxl formulas
(LibreOffice n'est pas utilisé : il est indisponible dans certains
environnements de calcul, et l'évaluation pure Python suffit ici.)
"""

import glob
import os
import re
import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile

import formulas
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "mise_en_donnees_materiaux.xlsx")
MIXDIR = os.path.join(HERE, "..", "data", "mixtures")
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", NS)

# onglet -> (fichier XML, nom de la composition du gaz de pyrolyse)
LINKS = {
    "TACOT":            ("tacot-air_25.xml",  "tacot_pyro"),
    "CPh70":            ("cph70-air_25.xml",  "cph70_pyro"),
    "ZURAM":            ("zuram-air.xml",     "VKIZuramPyroGas"),
    "SC-1008":          ("sc1008-air.xml",    "sc1008_pyro"),
    "Liège-phénolique": ("cork-air.xml",      "cork_pyro"),
}
TOL = 1e-3   # écart relatif toléré : l'arrondi des XML (3 à 4 décimales)


def xml_composition(fname, name):
    txt = open(os.path.join(MIXDIR, fname), encoding="utf-8").read()
    m = re.search(r'<composition name="%s"[^>]*>([^<]+)<' % re.escape(name), txt)
    out = {}
    for el, val in re.findall(r"([A-Za-z]+)\s*:\s*([0-9.eE+-]+)", m.group(1)):
        out[el] = float(val)
    return out


def evaluate(path):
    xl = formulas.ExcelModel().loads(path).finish()
    sol = xl.calculate()
    pat = re.compile(r"^'\[(?P<book>[^\]]+)\](?P<sheet>[^']+)'!(?P<cell>[A-Z]+\d+)$")
    vals = {}
    for k, v in sol.items():
        m = pat.match(k)
        if not m:
            continue
        try:
            vals[(m.group("sheet").upper(), m.group("cell"))] = v.value[0, 0]
        except Exception:
            pass
    return vals


def inject(path, names, vals):
    """Écrit les valeurs calculées dans le cache <v> de chaque cellule formule."""
    tmp = path + ".tmp"
    zin, zout = zipfile.ZipFile(path), zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    n = 0
    for item in zin.infolist():
        data = zin.read(item.filename)
        m = re.match(r"xl/worksheets/sheet(\d+)\.xml$", item.filename)
        if m:
            sheet = names[int(m.group(1)) - 1].upper()
            root = ET.fromstring(data)
            for c in root.iter("{%s}c" % NS):
                if c.find("{%s}f" % NS) is None:
                    continue
                v = vals.get((sheet, c.get("r")))
                if v is None or isinstance(v, (str, bool)):
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                old = c.find("{%s}v" % NS)
                if old is not None:
                    c.remove(old)
                ET.SubElement(c, "{%s}v" % NS).text = repr(fv)
                n += 1
            data = ET.tostring(root, encoding="UTF-8", xml_declaration=True)
        zout.writestr(item, data)
    zin.close(); zout.close(); shutil.move(tmp, path)
    return n


def main():
    ko = 0
    vals = evaluate(XLSX)
    wb = load_workbook(XLSX)

    # --- 1. erreurs de formule ------------------------------------------------
    nf, errs = 0, []
    for ws in wb.worksheets:
        for line in ws.iter_rows():
            for c in line:
                if isinstance(c.value, str) and c.value.startswith("="):
                    nf += 1
                    v = vals.get((ws.title.upper(), c.coordinate), "<non évaluée>")
                    if isinstance(v, str):
                        errs.append(f"{ws.title}!{c.coordinate} -> {v}   {c.value}")
    print("=" * 76)
    print("  1. ÉVALUATION DES FORMULES")
    print("=" * 76)
    print(f"  {nf} formules évaluées, {len(errs)} erreur(s)")
    for e in errs[:20]:
        print("   ", e)
    ko += len(errs)

    # --- 2. comparaison aux fichiers XML du dépôt -----------------------------
    print("\n" + "=" * 76)
    print("  2. CHAÎNE DE CALCUL  vs  data/mixtures/*.xml")
    print("=" * 76)
    for sheet, (fname, comp) in LINKS.items():
        ws = wb[sheet]
        rr = next(c.row for c in ws["B"]
                  if isinstance(c.value, str) and "valeur portée dans le XML" in c.value)
        calc = {e: vals[(sheet.upper(), f"{col}{rr-1}")]
                for e, col in zip("CHO", "CDE")}
        ref = xml_composition(fname, comp)
        tot = sum(ref.values())
        ref = {e: v / tot for e, v in ref.items()}
        worst = max(abs(calc[e] - ref[e]) / ref[e] for e in "CHO")
        flag = "OK " if worst < TOL else "KO "
        ko += 0 if worst < TOL else 1
        print(f"  [{flag}] {sheet:18s} {comp:16s} "
              f"calcul C={calc['C']:.4f} H={calc['H']:.4f} O={calc['O']:.4f} | "
              f"XML C={ref['C']:.4f} H={ref['H']:.4f} O={ref['O']:.4f} | "
              f"écart max {worst*100:.2f} %")
    for sheet, fname, comp, exp in [("Carbone", "carbon-air.xml", "pyro", {"C": 1.0}),
                                    ("Silice", "silica-air.xml", "silice", {"Si": 1.0, "O": 2.0})]:
        got = xml_composition(fname, comp)
        good = got == exp
        ko += 0 if good else 1
        print(f"  [{'OK ' if good else 'KO '}] {sheet:18s} {comp:16s} "
              f"stœchiométrie directe : {got}")

    # --- 3. injection des valeurs --------------------------------------------
    n = inject(XLSX, [w.title for w in wb.worksheets], vals)
    print("\n" + "=" * 76)
    print(f"  3. VALEURS INJECTÉES DANS LE CACHE : {n}")
    print("=" * 76)
    print("  [OK] classeur vérifié" if ko == 0 else f"  [KO] {ko} problème(s)")
    return 1 if ko else 0


if __name__ == "__main__":
    sys.exit(main())
