#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Construit `mise_en_donnees_materiaux.xlsx` : un onglet par materiau etudie,
detaillant les hypotheses et TOUTES les etapes de calcul (formules vivantes)
qui menent aux proportions portees dans les fichiers `data/mixtures/*.xml`.

Trame commune a chaque onglet materiau :

    SS0 - Fiche d'identite
    SS1 - Hypotheses
    SS2 - Donnees d'entree (avec sources)
    SS3 - Calcul pas a pas
    SS4 - Resultat : ce qui entre dans le XML
    SS5 - Controles et validation croisee
    SS6 - Sensibilite
    SS7 - Reponse materiau (hors XML)
    SS8 - Bloc XML

Toutes les cellules de resultat sont des FORMULES : changer une donnee d'entree
(bleue) recalcule la chaine jusqu'a la composition du XML.

Usage :
    python build_mise_en_donnees.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "mise_en_donnees_materiaux.xlsx")

# ---------------------------------------------------------------------------
# Chartre graphique
# ---------------------------------------------------------------------------
FONT = "Arial"

F_TITLE   = Font(name=FONT, size=16, bold=True, color="FFFFFF")
F_SUB     = Font(name=FONT, size=10, italic=True, color="FFFFFF")
F_SECT    = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_STEP    = Font(name=FONT, size=10, bold=True, color="1F3864")
F_HEAD    = Font(name=FONT, size=9,  bold=True)
F_LAB     = Font(name=FONT, size=10)
F_LABB    = Font(name=FONT, size=10, bold=True)
F_IN      = Font(name=FONT, size=10, color="0000FF")            # donnee source
F_CALC    = Font(name=FONT, size=10, color="000000")            # formule
F_LINK    = Font(name=FONT, size=10, color="008000")            # lien inter-onglet
F_OUT     = Font(name=FONT, size=10, bold=True, color="C00000")  # valeur du XML
F_NOTE    = Font(name=FONT, size=9,  italic=True, color="595959")
F_MONO    = Font(name="Courier New", size=9)

FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_SECT  = PatternFill("solid", fgColor="2E75B6")
FILL_HEAD  = PatternFill("solid", fgColor="D9E2F3")
FILL_KEY   = PatternFill("solid", fgColor="FFF2CC")   # hypothese cle / a remplir
FILL_OUT   = PatternFill("solid", fgColor="FCE4E4")   # resultat XML
FILL_XML   = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NF_X   = "0.0000"     # fraction molaire / massique
NF_X3  = "0.000"
NF_P   = "0.00"       # pourcentage exprime en points (% masse)
NF_MOL = "0.0000"
NF_M   = "0.000"
NF_E   = "0.00%"

WIDTHS = {"A": 4.5, "B": 34, "C": 13, "D": 13, "E": 13,
          "F": 13, "G": 13, "H": 13, "I": 72}


# ---------------------------------------------------------------------------
# Briques de mise en page
# ---------------------------------------------------------------------------
def setup(ws, tab_color="1F3864"):
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def title(ws, name, subtitle, xmlfiles):
    ws.merge_cells("A1:I1")
    ws["A1"] = name
    ws["A1"].font = F_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:I2")
    ws["A2"] = subtitle
    ws["A2"].font = F_SUB
    ws["A2"].fill = FILL_TITLE
    ws["A2"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 16
    ws.merge_cells("A3:I3")
    ws["A3"] = "Fichier(s) XML : " + xmlfiles
    ws["A3"].font = F_MONO
    ws["A3"].alignment = Alignment(vertical="center", indent=1)
    return 5


def section(ws, r, text):
    ws.merge_cells(f"A{r}:I{r}")
    ws[f"A{r}"] = text
    ws[f"A{r}"].font = F_SECT
    ws[f"A{r}"].fill = FILL_SECT
    ws[f"A{r}"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 18
    return r + 1


def step(ws, r, text, math=None):
    ws[f"A{r}"] = ">"
    ws[f"A{r}"].font = F_STEP
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = text
    ws[f"B{r}"].font = F_STEP
    if math:
        ws[f"I{r}"] = math
        ws[f"I{r}"].font = F_MONO
        ws[f"I{r}"].alignment = Alignment(vertical="center", wrap_text=True)
    return r + 1


def note(ws, r, text, col="B", span="I"):
    ws.merge_cells(f"{col}{r}:{span}{r}")
    ws[f"{col}{r}"] = text
    ws[f"{col}{r}"].font = F_NOTE
    ws[f"{col}{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    return r + 1


def para(ws, r, text, height=None):
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = text
    ws[f"B{r}"].font = F_LAB
    ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    if height:
        ws.row_dimensions[r].height = height
    return r + 1


def kv(ws, r, label, value, src="", font=F_IN, nf=None, tag=None, key=False):
    """Ligne 'libelle | valeur | source'."""
    if tag is not None:
        ws[f"A{r}"] = tag
        ws[f"A{r}"].font = F_HEAD
    ws[f"B{r}"] = label
    ws[f"B{r}"].font = F_LAB
    ws[f"C{r}"] = value
    ws[f"C{r}"].font = font
    if nf:
        ws[f"C{r}"].number_format = nf
    if key:
        ws[f"C{r}"].fill = FILL_KEY
    ws[f"I{r}"] = src
    ws[f"I{r}"].font = F_NOTE
    ws[f"I{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    return r + 1


def headrow(ws, r, labels, start="B"):
    c0 = ord(start) - 64
    for i, lab in enumerate(labels):
        c = get_column_letter(c0 + i)
        ws[f"{c}{r}"] = lab
        ws[f"{c}{r}"].font = F_HEAD
        ws[f"{c}{r}"].fill = FILL_HEAD
        ws[f"{c}{r}"].border = BOX
        ws[f"{c}{r}"].alignment = Alignment(horizontal="center", wrap_text=True)
    return r + 1


def row(ws, r, values, start="B", fonts=None, nfs=None, fills=None):
    c0 = ord(start) - 64
    for i, v in enumerate(values):
        c = get_column_letter(c0 + i)
        cell = ws[f"{c}{r}"]
        cell.value = v
        cell.font = (fonts[i] if fonts and fonts[i] else (F_LAB if i == 0 else F_CALC))
        cell.border = BOX
        if nfs and i < len(nfs) and nfs[i]:
            cell.number_format = nfs[i]
        if fills and i < len(fills) and fills[i]:
            cell.fill = fills[i]
    return r + 1


def xmlblock(ws, r, lines):
    for ln in lines:
        ws.merge_cells(f"B{r}:I{r}")
        ws[f"B{r}"] = ln
        ws[f"B{r}"].font = F_MONO
        ws[f"B{r}"].fill = FILL_XML
        ws[f"B{r}"].alignment = Alignment(vertical="center", indent=1)
        r += 1
    return r


def result_block(ws, r, name, cC, cH, cO, ref, labels=("C", "H", "O")):
    """Bloc résultat standard : x_E calculé, valeur du XML, écart."""
    r = headrow(ws, r, ["Composition élémentaire (fractions molaires)"] + list(labels) + ["Somme"])
    r = row(ws, r, ["calcul — chaîne complète ci-dessus", f"={cC}", f"={cH}", f"={cO}",
                    f"=SUM(C{r}:E{r})"],
            nfs=[None, NF_X, NF_X, NF_X, NF_X3])
    rcalc = r - 1
    r = row(ws, r, [f"valeur portée dans le XML ({name})", ref[0], ref[1], ref[2],
                    f"=SUM(C{r}:E{r})"],
            fonts=[F_LAB, F_OUT, F_OUT, F_OUT, F_CALC],
            nfs=[None, NF_X, NF_X, NF_X, NF_X3],
            fills=[None, FILL_OUT, FILL_OUT, FILL_OUT, None])
    rxml = r - 1
    r = row(ws, r, ["écart relatif |calcul − XML| / XML",
                    f"=ABS(C{rcalc}-C{rxml})/C{rxml}",
                    f"=ABS(D{rcalc}-D{rxml})/D{rxml}",
                    f"=ABS(E{rcalc}-E{rxml})/E{rxml}", ""],
            nfs=[None, NF_E, NF_E, NF_E, None])
    return r, rcalc, rxml


# ---------------------------------------------------------------------------
# Classeur
# ---------------------------------------------------------------------------
wb = Workbook()
ws_syn = wb.active
ws_syn.title = "Synthèse"
ws_cst = wb.create_sheet("Constantes")
ws_tac = wb.create_sheet("TACOT")
ws_cph = wb.create_sheet("CPh70")
ws_zur = wb.create_sheet("ZURAM")
ws_sc  = wb.create_sheet("SC-1008")
ws_cor = wb.create_sheet("Liège-phénolique")
ws_car = wb.create_sheet("Carbone")
ws_sil = wb.create_sheet("Silice")

# Adresses absolues des masses molaires (onglet Constantes)
MC  = "Constantes!$C$7"    # C   usuelle
MH  = "Constantes!$C$8"    # H
MO  = "Constantes!$C$9"    # O
MN  = "Constantes!$C$10"   # N
MSI = "Constantes!$C$11"   # Si
MCz = "Constantes!$D$7"    # C  masses du classeur ZURAM
MHz = "Constantes!$D$8"
MOz = "Constantes!$D$9"


# ===========================================================================
# ONGLET  CONSTANTES
# ===========================================================================
ws = ws_cst
setup(ws, "808080")
r = title(ws, "Constantes et conventions communes",
          "Masses molaires, air de couche limite, motifs de résine, bilan de masse de surface — référencés par tous les onglets matériaux",
          "toutes les mises en données de data/mixtures/")

r = section(ws, r, "§1 — MASSES MOLAIRES DES ÉLÉMENTS")
r = headrow(ws, r, ["Élément", "M usuelle [g/mol]", "M classeur ZURAM [g/mol]", "", ""])
r = row(ws, r, ["Carbone      C",  12.011,  12.0107, "", ""],
        fonts=[F_LAB, F_IN, F_IN, None, None], nfs=[None, "0.0000", "0.0000"])
r = row(ws, r, ["Hydrogène    H", 1.008,   1.00794, "", ""],
        fonts=[F_LAB, F_IN, F_IN, None, None], nfs=[None, "0.00000", "0.00000"])
r = row(ws, r, ["Oxygène      O", 15.999,  15.999,  "", ""],
        fonts=[F_LAB, F_IN, F_IN, None, None], nfs=[None, "0.0000", "0.0000"])
r = row(ws, r, ["Azote        N", 14.007,  "—",     "", ""],
        fonts=[F_LAB, F_IN, F_NOTE, None, None], nfs=[None, "0.0000", None])
r = row(ws, r, ["Silicium     Si", 28.0855, "—",    "", ""],
        fonts=[F_LAB, F_IN, F_NOTE, None, None], nfs=[None, "0.0000", None])
r = note(ws, r, "Colonne « usuelle » : valeurs employées par les scripts du dépôt "
                "(cork_pyrolysis_data.py). Colonne « classeur ZURAM » : valeurs de "
                "Tacot_Zuram_Calcarb_database_v4.3.1.ods, onglet ZURAM_official!F87:F89 — "
                "elles seules reproduisent les fractions molaires de ce classeur au 1e-16 près.")
r += 1

r = section(ws, r, "§2 — BORD DE COUCHE LIMITE : L'AIR (identique dans tous les XML)")
rair = r + 1
r = headrow(ws, r, ["Air (option -bl)", "N", "O", "Somme", ""])
r = row(ws, r, ["fraction molaire élémentaire x_E  (XML)", 0.79, 0.21, f"=SUM(C{r}:D{r})", ""],
        fonts=[F_LAB, F_OUT, F_OUT, F_CALC, None], nfs=[None, NF_X3, NF_X3, NF_X3])
r = row(ws, r, ["masse molaire M_E [g/mol]", f"={MN}", f"={MO}", "", ""],
        fonts=[F_LAB, F_LINK, F_LINK, None, None], nfs=[None, NF_M, NF_M])
r = row(ws, r, ["x_E · M_E", f"=C{rair}*C{rair+1}", f"=D{rair}*D{rair+1}",
                f"=SUM(C{r}:D{r})", ""], nfs=[None, NF_M, NF_M, NF_M])
r = row(ws, r, ["fraction massique élémentaire y_E", f"=C{rair+2}/$E${rair+2}",
                f"=D{rair+2}/$E${rair+2}", f"=SUM(C{r}:D{r})", ""],
        nfs=[None, NF_X, NF_X, NF_X3])
r = note(ws, r, "y_E = x_E·M_E / Σ x_j·M_j.  Le parseur de Mutation++ normalise seul la somme à 1 "
                "et convertit mole ↔ masse (Mixture.cpp:131-138) : `type=\"mole\"` est le défaut, "
                "`type=\"mass\"` doit être déclaré explicitement si la source est massique.")
r += 1

r = section(ws, r, "§3 — MOTIFS DE RÉSINE PHÉNOLIQUE (référence commune TACOT / ZURAM / SC-1008 / liège)")
r = step(ws, r, "Un motif cuit s'écrit C(6+b)H6O : un cycle phénolique et b ponts méthylène –CH2–.",
         "M = ν_C·M_C + ν_H·M_H + ν_O·M_O ;  w_E = ν_E·M_E / M")
rm = r + 1
r = headrow(ws, r, ["Motif", "ν_C", "ν_H", "ν_O", "% C masse", "% H masse", "% O masse"])
for lab, nc, nh, no, com in [("phénol C6H6O (monomère)", 6, 6, 1, "TACOT : motif équivalent mesuré (Sykes table II), à 0.85 point près"),
                             ("novolac linéaire C7H6O (b = 1)", 7, 6, 1, "ZURAM : motif équivalent mesuré ([THo] fig. 7), à 0.35 point près ; liège : motif retenu pour la résine"),
                             ("résol saturé / réticulé C7.5H6O (b = 1.5)", 7.5, 6, 1, "SC-1008 : motif RETENU (résol saturé) — M = 112.13 g/mol")]:
    r = row(ws, r, [lab, nc, nh, no,
                    f"=C{r}*{MC}/(C{r}*{MC}+D{r}*{MH}+E{r}*{MO})*100",
                    f"=D{r}*{MH}/(C{r}*{MC}+D{r}*{MH}+E{r}*{MO})*100",
                    f"=E{r}*{MO}/(C{r}*{MC}+D{r}*{MH}+E{r}*{MO})*100"],
            fonts=[F_LAB, F_IN, F_IN, F_IN, F_CALC, F_CALC, F_CALC],
            nfs=[None, NF_X3, NF_X3, NF_X3, NF_P, NF_P, NF_P])
    ws[f"I{r-1}"] = com
    ws[f"I{r-1}"].font = F_NOTE
    ws[f"I{r-1}"].alignment = Alignment(vertical="top", wrap_text=True)
r = note(ws, r, "Le nombre de ponts b n'est pas un réglage libre : le phénol a 3 sites réactifs "
                "(2 ortho + 1 para) et un pont est partagé entre 2 cycles, d'où b ≤ 1.5 à réseau saturé. "
                "Novolac (formé en défaut de formaldéhyde, durci à l'hexamine) et résol (excès de "
                "formaldéhyde, auto-réticulant) glissent sur la MÊME courbe : seul b diffère. "
                "L'analyse élémentaire C/H/O ne permet donc pas de distinguer les deux — seul l'azote le fait.")
r += 1

r = section(ws, r, "§4 — LE BILAN QUI CONSOMME CES COMPOSITIONS (rappel)")
r = para(ws, r, "Bilan de conservation de chaque élément i à la paroi, couche limite laminaire à nombre de Lewis unité :", 16)
r = xmlblock(ws, r, [
    "   (1 + B'c + B'g) · Y_w,i  =  Y_e,i  +  B'c · Y_c,i  +  B'g · Y_g,i",
    "",
    "   =>   B'c = [ Y_e,i + B'g·Y_g,i − Y_w,i·(1 + B'g) ] / ( Y_w,i − Y_c,i )",
])
r = para(ws, r, "Y_e = bord de couche limite (option -bl) · Y_g = gaz de pyrolyse (-py) · "
                "Y_c = char (-char, avec -char-elem) · Y_w = gaz à la paroi, calculé à l'équilibre "
                "thermochimique à (T_w, P). `bprime` ne consomme QUE ces trois compositions "
                "élémentaires, plus T, P et B'g (src/apps/bprime.cpp:304-330).", 30)
r = para(ws, r, "Conséquence majeure : aucune grandeur volumique (masse volumique, porosité, "
                "rapport fibres/résine, rendement en char, cinétique) n'entre dans le XML. Toutes "
                "vivent dans le solveur de réponse matériau — voir le §7 de chaque onglet.", 30)
r += 1

r = section(ws, r, "§5 — LÉGENDE DES COULEURS ET MODE D'EMPLOI")
for txt, fnt, fill in [
        ("bleu — donnée d'entrée en dur : mesure, valeur publiée, hypothèse. C'est ici qu'on édite.", F_IN, None),
        ("noir — cellule calculée par formule vivante. Ne pas écraser.", F_CALC, None),
        ("vert — lien vers un autre onglet (constantes, ou composition héritée d'un autre matériau).", F_LINK, None),
        ("rouge sur fond rosé — la valeur telle qu'elle figure dans le fichier XML du dépôt.", F_OUT, FILL_OUT),
        ("fond jaune — hypothèse clé : c'est le paramètre dont dépend le plus le résultat.", F_LAB, FILL_KEY),
]:
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = txt
    ws[f"B{r}"].font = fnt
    if fill:
        ws[f"B{r}"].fill = fill
    r += 1
r += 1
r = para(ws, r, "Chaque onglet matériau suit la même trame : §0 fiche d'identité · §1 hypothèses · "
                "§2 données d'entrée · §3 calcul pas à pas · §4 résultat porté dans le XML · "
                "§5 contrôles et validation croisée · §6 sensibilité · §7 réponse matériau (hors XML) · "
                "§8 bloc XML.", 30)


def hyp(ws, r, tag, text, why=""):
    ws[f"A{r}"] = tag
    ws[f"A{r}"].font = F_STEP
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = text
    ws[f"B{r}"].font = F_LAB
    ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    ws[f"I{r}"] = why
    ws[f"I{r}"].font = F_NOTE
    ws[f"I{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 26
    return r + 1


# ===========================================================================
# ONGLET  TACOT
# ===========================================================================
ws = ws_tac
setup(ws, "C00000")
r = title(ws, "TACOT — Theoretical Ablative Composite for Open Testing",
          "Carbone/phénolique poreux · gaz de pyrolyse obtenu par SOMMATION ATOMIQUE d'une spéciation moléculaire mesurée",
          "data/mixtures/tacot-air_25.xml · tacot-air_35.xml · tacot-pyrogas.xml")

r = section(ws, r, "§0 — FICHE D'IDENTITÉ")
r = kv(ws, r, "Matériau", "TACOT 3.0 (matériau de référence ouvert)", "Classeur TACOT_3.0.xls, onglet Description")
r = kv(ws, r, "Renfort", "fibres de carbone ex-cellulose, traitées à 2000 K, 1600 kg/m³", "Description!A11 — « The carbon fibers do not pyrolyse » ('Pyrolysis model'!H20)")
r = kv(ws, r, "Résine", "novolac phénol-formaldéhyde Union Carbide BRP 5549 + hexamine (HMTA), 1200 kg/m³", "Sykes, NASA TN D-3810 (1967), § MATERIAL p. 2 ; Description!A12")
r = kv(ws, r, "Microstructure", "ε_fibres 0.10 / ε_matrice 0.10 / porosité 0.80", "Description!A15:A17")
r = kv(ws, r, "Source primaire du gaz", "Sykes, table I (chromatographie, paliers de 50 °C entre 100 et 1000 °C)", "reprise dans 'Pyrolysis-gas chemistry'!B5:H5 puis 'Pyrolysis model'!B4:B6")
r = kv(ws, r, "Compositions du XML", "air (-bl) · tacot_pyro (-py) · tacot_char (-char, -char-elem C)", "tacot_bprime/resine_tacot.md et mise_en_donnees_xml.md")
r += 1

r = section(ws, r, "§1 — HYPOTHÈSES")
r = hyp(ws, r, "H1", "Le gaz de pyrolyse est produit par la RÉSINE SEULE.",
        "Les fibres ex-cellulose sont déjà traitées à 2000 K : elles ne pyrolysent pas ('Pyrolysis model'!H20). Le rapport fibres/résine ne fixe donc que la QUANTITÉ de gaz (donc B'g), pas sa composition — cf. onglet CPh70.")
r = hyp(ws, r, "H2", "La spéciation retenue est la somme RÉELLE des colonnes de la table I de Sykes (100.30), pas les totaux imprimés arrondis (100.0).",
        "C'est ce choix, et lui seul, qui permet de reproduire le classeur TACOT 3.0 au 1e-16 près (§5).")
r = hyp(ws, r, "H3", "Réduction du modèle : toluène C7H8 → benzène C6H6, et 2,4-xylénol C8H10O → phénol C6H5OH.",
        "'Pyrolysis-gas chemistry'!A8 : compatibilité avec le mécanisme cinétique d'April (10 réactions), qui ignore toluène et xylénol. Conservative en atomes à ~1 % près (C7H8→C6H6 perd un CH2).")
r = hyp(ws, r, "H4", "Le gaz est déclaré SANS AZOTE.",
        "Les 2.35 % de N de la résine viennent de la HMTA prémélangée ; ils s'effondrent à 0.42 % dès 400 °C et à zéro à 500 °C (Sykes table II) : signature d'un composé piégé, pas d'un azote de structure.")
r = hyp(ws, r, "H5", "Le char est du carbone pur : C:1.0.",
        "Sykes mesure pourtant C 92.65 / H 0.90 / O 6.45 % à 850 °C. L'hypothèse reste licite en ablation : la paroi dépasse 2000 K (l'oxygène résiduel part en CO bien avant) et les fibres, 57 % de la masse vierge, sont déjà du carbone pur.")
r = hyp(ws, r, "H6", "Le gaz à la paroi est à l'équilibre thermochimique à (T_w, P).",
        "Hypothèse centrale de la table B'. Le solveur multiphase de Mutation++ traite la coexistence gaz / C(gr).")
r = hyp(ws, r, "H7", "Liste d'espèces = les 25 espèces gazeuses de la table B' de référence, PLUS la phase condensée C(gr).",
        "Sans C(gr) l'excès de carbone ne peut pas se condenser : Y_w,C → 1 et B'c sature vers ≈ 200 au lieu de ≈ 0.087 à 300 K. La liste conditionne le résultat : 25 espèces donnent 0.43 % d'écart moyen contre la table de référence, 35 espèces 2.41 %.")
r += 1

r = section(ws, r, "§2 — DONNÉES D'ENTRÉE")
r = step(ws, r, "Table I de Sykes — distribution des produits de décomposition, en % des moles totales observées",
         "somme sur les 19 paliers de température")
rt1 = r + 1
r = headrow(ws, r, ["Espèce (table I)", "Σ colonne (réelle)", "total imprimé", "après réduction H3", "", "", ""])
sykes = [("CO2", 1.57, 1.6, "CO2"), ("CO", 5.78, 5.5, "CO"),
         ("C6H6  benzène", 0.17, 0.2, "C6H6"), ("C7H8  toluène", 0.31, 0.3, "→ C6H6"),
         ("C6H5OH  phénol", 7.14, 7.1, "C6H5OH"), ("(CH3)2C6H3OH  xylénol", 1.80, 1.8, "→ C6H5OH"),
         ("CH4", 10.03, 10.0, "CH4"), ("H2O", 23.43, 23.4, "H2O"), ("H2", 50.07, 50.1, "H2")]
for lab, real, printed, red in sykes:
    r = row(ws, r, [lab, real, printed, red, "", "", ""],
            fonts=[F_LAB, F_IN, F_IN, F_NOTE, None, None, None],
            nfs=[None, NF_P, NF_P, None])
rt1_end = r - 1
r = row(ws, r, ["SOMME", f"=SUM(C{rt1}:C{rt1_end})", f"=SUM(D{rt1}:D{rt1_end})",
                "← 100.30 contre 100.0 : voir H2", "", "", ""],
        fonts=[F_LABB, F_CALC, F_CALC, F_NOTE, None, None, None],
        nfs=[None, NF_P, NF_P, None])
rtot = r - 1
r = note(ws, r, "Les totaux imprimés au bas de la table de Sykes sont arrondis et somment à 100.0 ; "
                "les colonnes elles-mêmes somment à 100.30. C'est cette dernière valeur que le classeur "
                "TACOT 3.0 utilise pour renormaliser.")
r += 1
r = kv(ws, r, "Analyse élémentaire de la résine vierge — C [% masse]", 75.60, "Sykes, table II, ligne « ambiant ». L'analyse ne dose que C, H et N ; O est le complément à 100 % (cendres mesurées à 0.04 %, négligeables).")
rres = r - 1
r = kv(ws, r, "                                        H [% masse]", 6.12, "")
r = kv(ws, r, "                                        N [% masse]", 2.35, "issu de la HMTA — retiré du gaz, cf. H4")
r = kv(ws, r, "                                        O [% masse]", 15.93, "complément à 100 %")
r = kv(ws, r, "Rendement en char de la résine — mesuré", 0.540, "Sykes table II : fraction massique résiduelle à 850 °C", key=True)
ry_mes = r - 1
r = kv(ws, r, "Rendement en char de la résine — modèle TACOT", 0.500, "implicite : microstructure ('Pyrolysis model'!A32:A34) ET cinétique de Goldstein (B18:C19) donnent 0.50 toutes deux", key=True)
ry_mod = r - 1
r = kv(ws, r, "ρ fibres / ρ résine [kg/m³]", 1600, "Description!A11")
rrhof = r - 1
r = kv(ws, r, "", 1200, "Description!A12")
rrhom = r - 1
r = kv(ws, r, "ε_fibres / ε_matrice [-]", 0.10, "Description!A15:A17 (porosité 0.80)")
repsf = r - 1
r = kv(ws, r, "", 0.10, "")
repsm = r - 1
r += 1

r = section(ws, r, "§3 — CALCUL PAS À PAS")
r = step(ws, r, "Étape 1 — regrouper les 9 espèces de Sykes en 7 (réduction H3), puis renormaliser par la somme réelle",
         "x_j = n_j / Σ n_j     avec Σ n_j = 100.30")
rsp = r + 1
r = headrow(ws, r, ["Espèce du modèle", "moles brutes", "x_j calculé", "classeur TACOT 3.0",
                    "ν_C", "ν_H", "ν_O"])
red = [("CO2",     f"=C{rt1}",                  0.0156530408773679, 1, 0, 2),
       ("CO",      f"=C{rt1+1}",                0.0576271186440678, 1, 0, 1),
       ("C6H6",    f"=C{rt1+2}+C{rt1+3}",       0.0047856430707876, 6, 6, 0),
       ("C6H5OH",  f"=C{rt1+4}+C{rt1+5}",       0.0891326021934197, 6, 6, 1),
       ("CH4",     f"=C{rt1+6}",                0.1000000000000000, 1, 4, 0),
       ("H2O",     f"=C{rt1+7}",                0.2335992023928220, 0, 2, 1),
       ("H2",      f"=C{rt1+8}",                0.4992023928215351, 0, 2, 0)]
for lab, f, ref, nc, nh, no in red:
    r = row(ws, r, [lab, f, f"=C{r}/$C${rtot}", ref, nc, nh, no],
            fonts=[F_LAB, F_CALC, F_CALC, F_IN, F_IN, F_IN, F_IN],
            nfs=[None, NF_P, "0.0000000000", "0.0000000000", "0", "0", "0"])
rsp_end = r - 1
r = row(ws, r, ["SOMME", f"=SUM(C{rsp}:C{rsp_end})", f"=SUM(D{rsp}:D{rsp_end})",
                f"=SUM(E{rsp}:E{rsp_end})", "", "", ""],
        fonts=[F_LABB, F_CALC, F_CALC, F_CALC, None, None, None],
        nfs=[None, NF_P, NF_X3, NF_X3])
_terms = ",".join(f"ABS(D{i}-E{i})" for i in range(rsp, rsp_end + 1))
r = kv(ws, r, "écart maximal au classeur, espèce par espèce", f"=MAX({_terms})",
       "Colonne « classeur TACOT 3.0 » : 'Pyrolysis-gas chemistry'!B5:H5. Attendu ~5e-16, "
       "la précision de la double précision : la chaîne Sykes → classeur est entièrement "
       "élucidée.", font=F_CALC, nf="0.00E+00")
rsp_tot = r - 1
r += 1

r = step(ws, r, "Étape 2 — sommation atomique : passer du moléculaire à l'élémentaire",
         "n_E = Σ_j ν_E,j · x_j   (SOMMEPROD)")
ratom = r + 1
r = headrow(ws, r, ["Moles d'atomes par mole de gaz", "C", "H", "O", "Somme", "", ""])
r = row(ws, r, ["n_E = Σ_j ν_E,j · x_j",
                f"=SUMPRODUCT($D${rsp}:$D${rsp_end},F{rsp}:F{rsp_end})",
                f"=SUMPRODUCT($D${rsp}:$D${rsp_end},G{rsp}:G{rsp_end})",
                f"=SUMPRODUCT($D${rsp}:$D${rsp_end},H{rsp}:H{rsp_end})",
                f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rn = r - 1
r = step(ws, r, "Étape 3 — normaliser en fractions molaires élémentaires, puis convertir en fractions massiques",
         "x_E = n_E / Σ n_j ;  M_gaz = Σ x_E·M_E ;  y_E = x_E·M_E / M_gaz")
rx = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "M_gaz [g/mol]", ""])
r = row(ws, r, ["x_E (fraction molaire élémentaire)", f"=C{rn}/$F${rn}", f"=D{rn}/$F${rn}",
                f"=E{rn}/$F${rn}", f"=SUM(C{r}:E{r})",
                f"=C{r}*{MC}+D{r}*{MH}+E{r}*{MO}", ""],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3, NF_M])
rxE = r - 1
r = row(ws, r, ["y_E (fraction massique élémentaire)", f"=C{rxE}*{MC}/$G${rxE}",
                f"=D{rxE}*{MH}/$G${rxE}", f"=E{rxE}*{MO}/$G${rxE}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3])
r = note(ws, r, "M_gaz ≈ 5.0 g/mol : un gaz très léger, dominé par H2 (50 % des moles). "
                "En masse le gaz est pourtant à ~49 % de carbone.")
r += 1

r = section(ws, r, "§4 — RÉSULTAT : CE QUI ENTRE DANS LE XML")
r, rcalc, rxml = result_block(ws, r, "tacot_pyro", f"C{rxE}", f"D{rxE}", f"E{rxE}",
                              (0.206, 0.679, 0.115))
r = note(ws, r, "L'écart est celui de l'arrondi à trois décimales opéré en 'Pyrolysis model'!B4:B6. "
                "Mutation++ normalise la somme à 1 : porter C:0.7368, H:2.4291, O:0.4117 (les moles brutes) "
                "donnerait rigoureusement le même mélange.")
r += 1
r = headrow(ws, r, ["Char (tacot_char)", "C", "H", "O", "", "", ""])
r = row(ws, r, ["composition du XML", 1.0, "—", "—", "", "", ""],
        fonts=[F_LAB, F_OUT, F_NOTE, F_NOTE, None, None, None],
        nfs=[None, NF_X3, None, None], fills=[None, FILL_OUT, None, None])
r = note(ws, r, "Fibres de carbone et résine carbonisée donnent toutes deux du carbone pur : "
                "C:1.0 quel que soit le rapport fibres/résine (cf. H5). Option `-char-elem C`.")
r += 1

r = section(ws, r, "§5 — CONTRÔLES ET VALIDATION CROISÉE")
r = step(ws, r, "Contrôle A — reproduction du classeur TACOT 3.0 : écart maximal sur les 7 fractions molaires",
         "cf. colonne F de l'étape 1")
r = kv(ws, r, "écart maximal espèce par espèce", f"=C{rsp_tot}",
       "reporté de l'étape 1 — attendu ≈ 5e-16 (double précision)",
       font=F_CALC, nf="0.00E+00")
r = step(ws, r, "Contrôle B — fermeture inverse INDÉPENDANTE : partir de la résine mesurée (table II), "
                "retrancher un char de carbone pur, comparer au gaz de la table I",
         "n_E(gaz) = n_E(résine) − n_E(char) ;  n_C(char) = Y·100 / M_C")
rfc = r + 1
r = headrow(ws, r, ["Base 100 g de résine vierge", "C", "H", "O", "Somme", "écart max / table I", ""])
r = row(ws, r, ["moles d'atomes de la résine (table II)",
                f"=C{rres}/{MC}", f"=C{rres+1}/{MH}", f"=C{rres+3}/{MO}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rmr = r - 1
r = row(ws, r, [f"char retranché, Y = 0.540 (mesuré)", f"=$C${ry_mes}*100/{MC}", 0, 0, "", "", ""],
        fonts=[F_LAB, F_CALC, F_CALC, F_CALC, None, None, None],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL])
rch = r - 1
r = row(ws, r, ["gaz obtenu — x_E", f"=(C{rmr}-C{rch})/($F${rmr}-$C${rch})",
                f"=(D{rmr}-D{rch})/($F${rmr}-$C${rch})",
                f"=(E{rmr}-E{rch})/($F${rmr}-$C${rch})", f"=SUM(C{r}:E{r})",
                f"=MAX(ABS(C{r}-C{rcalc})/C{rcalc},ABS(D{r}-D{rcalc})/D{rcalc},ABS(E{r}-E{rcalc})/E{rcalc})", ""],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3, NF_E])
rgy54 = r - 1
r = note(ws, r, "Accord à 2.4 % entre DEUX TECHNIQUES EXPÉRIMENTALES INDÉPENDANTES : analyse élémentaire "
                "du résidu (DTA, table II) d'un côté, chromatographie des effluents (table I) de l'autre. "
                "C'est la validation la plus forte de la cohérence interne de Sykes — et, accessoirement, "
                "du fait que le char soit assimilable à du carbone pur.")
r += 1

r = section(ws, r, "§6 — SENSIBILITÉ AU RENDEMENT EN CHAR")
rsens = r + 1
r = headrow(ws, r, ["Rendement en char Y", "C", "H", "O", "écart max / table I", "origine de la valeur", ""])
for Y, org in [(0.500, "modèle TACOT : microstructure + cinétique de Goldstein"),
               (0.534, "ajustement optimal (moindre écart)"),
               (0.540, "Sykes table II — mesuré à 850 °C")]:
    r = row(ws, r, [Y,
                    f"=(($C${rmr}-B{r}*100/{MC}))/($F${rmr}-B{r}*100/{MC})",
                    f"=$D${rmr}/($F${rmr}-B{r}*100/{MC})",
                    f"=$E${rmr}/($F${rmr}-B{r}*100/{MC})",
                    f"=MAX(ABS(C{r}-$C${rcalc})/$C${rcalc},ABS(D{r}-$D${rcalc})/$D${rcalc},ABS(E{r}-$E${rcalc})/$E${rcalc})",
                    org, ""],
            fonts=[F_IN, F_CALC, F_CALC, F_CALC, F_CALC, F_NOTE, None],
            nfs=[NF_X3, NF_X, NF_X, NF_X, NF_E, None])
r = note(ws, r, "Le rendement en char n'entre PAS dans le calcul du gaz du XML : celui-ci vient "
                "directement de la table I. Ce tableau ne sert qu'au contrôle croisé — et il montre que "
                "le 0.50 du modèle TACOT est une valeur arrondie pour produire des densités rondes "
                "(280/220 kg/m³), pas la mesure (0.540).")
r += 1

r = section(ws, r, "§7 — RÉPONSE MATÉRIAU (n'entre PAS dans le XML)")
r = step(ws, r, "Bilan de phase solide et couplage en ablation stationnaire",
         "ρ_v = ε_f·ρ_f + ε_m·ρ_m ;  ρ_c = ε_f·ρ_f + Y·ε_m·ρ_m ;  k = B'g/B'c = (ρ_v − ρ_c)/ρ_c")
r = kv(ws, r, "ρ vierge [kg/m³]", f"=$C${repsf}*$C${rrhof}+$C${repsm}*$C${rrhom}",
       "'Pyrolysis model'!A32 — attendu 280", font=F_CALC, nf="0.0")
rrv = r - 1
r = kv(ws, r, "ρ char [kg/m³]", f"=$C${repsf}*$C${rrhof}+$C${ry_mod}*$C${repsm}*$C${rrhom}",
       "'Pyrolysis model'!A33 — attendu 220 (avec Y = 0.50 du modèle)", font=F_CALC, nf="0.0")
rrc = r - 1
r = kv(ws, r, "gaz de pyrolyse libéré [kg/m³]", f"=C{rrv}-C{rrc}", "soit 21.4 % de la masse vierge",
       font=F_CALC, nf="0.0")
r = kv(ws, r, "couplage k = B'g / B'c", f"=(C{rrv}-C{rrc})/C{rrc}",
       "point de fonctionnement stationnaire : B'c = B'c_table(T, P, B'g = k·B'c)", font=F_CALC, nf="0.0000")
rk_tac = r - 1
r = note(ws, r, "k est la SEULE façon dont la microstructure entre dans le résultat final : elle ne change "
                "pas la table B'c(T, P, B'g), elle change l'endroit où le matériau la lit. Voir l'onglet CPh70, "
                "qui partage la table du TACOT mais pas son k.")
r += 1

r = section(ws, r, "§8 — BLOC XML")
r = xmlblock(ws, r, [
    '<mixture thermo_db="NASA-9">',
    '    <species>',
    '       C H O N CH4 CN CO CO2 C2 C2H C2H2,acetylene C3 C4 C4H2,butadiyne C5',
    '       HCN H2 H2O N2 CH2OH CNN CNC CNCOCN C6H6 HNC C(gr)',
    '    </species>',
    '    <element_compositions default="air">',
    '        <composition name="air">N:0.79, O:0.21</composition>',
    '        <composition name="tacot_pyro">C:0.206, H:0.679, O:0.115</composition>',
    '        <composition name="tacot_char">C:1.0</composition>',
    '    </element_compositions>',
    '</mixture>',
])
r += 1
r = para(ws, r, "bprime -T 300:25:4000 -P 101325 -b 0.5 -m tacot-air_25 -bl air "
                "-py tacot_pyro -char tacot_char -char-elem C", 16)
TAC_XML = (rxml, rk_tac, rxE, rrc)


# ===========================================================================
# ONGLET  CPh70
# ===========================================================================
ws = ws_cph
setup(ws, "ED7D31")
r = title(ws, "CPh70 — carbone/phénolique dense",
          "Mêmes fibres et même résine que le TACOT, autres proportions · démontre que le rapport fibres/résine N'ENTRE PAS dans le XML",
          "data/mixtures/cph70-air_25.xml")

r = section(ws, r, "§0 — FICHE D'IDENTITÉ")
r = kv(ws, r, "Matériau", "CPh70 — composite carbone/phénolique dense", "70 % fibres / 30 % résine en volume solide, porosité 0.01")
r = kv(ws, r, "Renfort", "fibres de carbone ex-cellulose, 1600 kg/m³ — IDENTIQUES au TACOT", "cf. onglet TACOT §0")
r = kv(ws, r, "Résine", "novolac/formaldéhyde, 1200 kg/m³ — IDENTIQUE au TACOT", "cf. onglet TACOT §0")
r = kv(ws, r, "Source du gaz", "héritée du TACOT (résine inchangée)", "tacot_bprime/cph70_vs_tacot.md")
r = kv(ws, r, "Compositions du XML", "air (-bl) · cph70_pyro (-py) · cph70_char (-char, -char-elem C)", "cph70_bprime/README.md")
r += 1

r = section(ws, r, "§1 — HYPOTHÈSES")
r = hyp(ws, r, "H1", "Fibres et résine RIGOUREUSEMENT identiques à celles du TACOT ; seuls changent les proportions et la porosité.",
        "C'est la définition même du cas d'étude. Si la résine ou la nature des fibres changeait, tout le §3 serait à refaire.")
r = hyp(ws, r, "H2", "Le gaz de pyrolyse est produit par la résine seule ⇒ sa composition ÉLÉMENTAIRE est inchangée.",
        "Changer le rapport fibres/résine change la QUANTITÉ de gaz produite (donc B'g), pas sa composition. Or B'g est un paramètre d'ENTRÉE de la table, balayé de 0 à 10.")
r = hyp(ws, r, "H3", "Fibres de carbone et résine carbonisée donnent toutes deux du carbone pur ⇒ char C:1.0 quelles que soient les proportions.",
        "La pondération générale n_E(char) = Σ_i (m_char,i/M_i)·ν_E,i se réduit à C:1.0 dès que tous les constituants carbonisent vers le même élément.")
r = hyp(ws, r, "H4", "Aucune grandeur volumique n'entre dans le bilan de surface.",
        "Le bilan est écrit sur des fractions massiques élémentaires, donc normalisées : ni ρ, ni porosité, ni ε_f n'y apparaissent (Constantes §4).")
r = hyp(ws, r, "H5", "Rendement en char de la résine : 0.50 (valeur du modèle TACOT).",
        "Elle n'entre pas dans le XML — seulement dans ρ_char, donc dans k (§7).")
r += 1

r = section(ws, r, "§2 — DONNÉES D'ENTRÉE")
r = kv(ws, r, "fraction volumique de fibres dans le SOLIDE", 0.70, "définition du CPh70", key=True)
rff = r - 1
r = kv(ws, r, "fraction volumique de résine dans le SOLIDE", 0.30, "complément à 1 − 0.70", key=True)
rfm = r - 1
r = kv(ws, r, "porosité φ", 0.01, "CPh70 : matériau dense (TACOT : 0.80)", key=True)
rphi = r - 1
r = kv(ws, r, "ρ intrinsèque fibres [kg/m³]", 1600, "Description!A11 du classeur TACOT 3.0")
rrhof = r - 1
r = kv(ws, r, "ρ intrinsèque résine [kg/m³]", 1200, "Description!A12")
rrhom = r - 1
r = kv(ws, r, "rendement en char de la résine", 0.50, "'Pyrolysis model'!A32:A34 et B18:C19", key=True)
ry = r - 1
r += 1

r = section(ws, r, "§3 — CALCUL PAS À PAS")
r = step(ws, r, "Étape 1 — composition du gaz de pyrolyse : elle est HÉRITÉE, pas recalculée",
         "Y_g(CPh70) = Y_g(TACOT), par H1 + H2")
rg = r + 1
r = headrow(ws, r, ["Gaz de pyrolyse", "C", "H", "O", "Somme", "", ""])
r = row(ws, r, ["reprise de l'onglet TACOT (chaîne complète : Sykes table I)",
                f"='TACOT'!C{TAC_XML[2]}", f"='TACOT'!D{TAC_XML[2]}", f"='TACOT'!E{TAC_XML[2]}",
                f"=SUM(C{r}:E{r})", "", ""],
        fonts=[F_LAB, F_LINK, F_LINK, F_LINK, F_CALC, None, None],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3])
rgc = r - 1
r = note(ws, r, "Toute la chaîne de traçabilité du gaz est celle de l'onglet TACOT : table I de Sykes → "
                "réduction toluène/xylénol → renormalisation par 100.30 → sommation atomique. "
                "Elle n'est PAS refaite ici : la résine est la même, donc le gaz est le même.")
r = step(ws, r, "Étape 2 — char : pondération générale des constituants",
         "n_E(char) = Σ_i (m_char,i / M_i) · ν_E,i")
r = para(ws, r, "Fibres : carbone pur (ν_C = 1, M = M_C). Résine carbonisée : carbone pur également. "
                "La somme pondérée de deux compositions identiques est cette composition, quelles que soient "
                "les masses : le char vaut C:1.0 pour TOUT rapport fibres/résine. C'est le point qui rendrait "
                "le calcul faux avec des fibres de silice ou un char retenant H/O.", 30)
r = step(ws, r, "Étape 3 — fractions volumiques réelles à partir de la porosité",
         "ε_i = (1 − φ) · f_i")
reps = r + 1
r = headrow(ws, r, ["Fractions volumiques", "fibres ε_f", "résine ε_m", "pores φ", "Somme", "", ""])
r = row(ws, r, ["ε = (1 − φ)·f",
                f"=(1-$C${rphi})*$C${rff}", f"=(1-$C${rphi})*$C${rfm}", f"=$C${rphi}",
                f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3])
repsr = r - 1
r = step(ws, r, "Étape 4 — bilan de phase solide : masses volumiques et couplage",
         "ρ_v = ε_f·ρ_f + ε_m·ρ_m ;  ρ_c = ε_f·ρ_f + Y·ε_m·ρ_m")
r = kv(ws, r, "ρ vierge [kg/m³]", f"=C{repsr}*$C${rrhof}+D{repsr}*$C${rrhom}",
       "attendu 1465.2 (TACOT : 280 — soit 5.2 fois plus dense)", font=F_CALC, nf="0.0")
rrv = r - 1
r = kv(ws, r, "ρ char [kg/m³]", f"=C{repsr}*$C${rrhof}+$C${ry}*D{repsr}*$C${rrhom}",
       "attendu 1287.0 (TACOT : 220)", font=F_CALC, nf="0.0")
rrc = r - 1
r = kv(ws, r, "fibres en % de la masse vierge", f"=C{repsr}*$C${rrhof}/C{rrv}*100",
       "attendu 75.7 % (TACOT : 57.1 %) — le CPh70 est un matériau riche en fibres", font=F_CALC, nf=NF_P)
r = kv(ws, r, "gaz de pyrolyse libéré [kg/m³]", f"=C{rrv}-C{rrc}",
       "attendu 178.2 kg/m³", font=F_CALC, nf="0.0")
r = kv(ws, r, "… soit en % de la masse vierge", f"=(C{rrv}-C{rrc})/C{rrv}*100",
       "12.2 % contre 21.4 % pour le TACOT : deux fois moins de gaz par unité de masse", font=F_CALC, nf=NF_P)
r = kv(ws, r, "couplage k = B'g / B'c = (ρ_v − ρ_c)/ρ_c", f"=(C{rrv}-C{rrc})/C{rrc}",
       "attendu 0.1385 — la SEULE grandeur qui distingue le CPh70 du TACOT", font=F_CALC, nf="0.0000")
rk = r - 1
r += 1

r = section(ws, r, "§4 — RÉSULTAT : CE QUI ENTRE DANS LE XML")
r, rcalc, rxml = result_block(ws, r, "cph70_pyro", f"C{rgc}", f"D{rgc}", f"E{rgc}",
                              (0.206, 0.679, 0.115))
r = headrow(ws, r, ["Char (cph70_char)", "C", "H", "O", "", "", ""])
r = row(ws, r, ["composition du XML", 1.0, "—", "—", "", "", ""],
        fonts=[F_LAB, F_OUT, F_NOTE, F_NOTE, None, None, None],
        nfs=[None, NF_X3, None, None], fills=[None, FILL_OUT, None, None])
r = note(ws, r, "Le fichier cph70-air_25.xml est donc STRICTEMENT identique à tacot-air_25.xml. "
                "Il n'est justifié que par la traçabilité — nommer le matériau. La table B'c(T, P, B'g) "
                "et h_w(T, P, B'g) sont rigoureusement les mêmes.")
r += 1

r = section(ws, r, "§5 — CONTRÔLES ET VALIDATION CROISÉE")
r = step(ws, r, "Contrôle A — écart mesuré entre les tables B' calculées CPh70 et TACOT", "sur B'c(T, B'g) aux 4 pressions")
r = kv(ws, r, "écart maximal", "0.000e+00", "vérifié numériquement — cf. tacot_bprime/cph70_vs_tacot.md §1", font=F_CALC)
r = step(ws, r, "Contrôle B — le B'c RÉELLEMENT atteint diffère, lui : c'est la droite B'g = k·B'c qui change",
         "point fixe : B'c = B'c_table(T, P, B'g = k·B'c)")
rpt = r + 1
r = headrow(ws, r, ["T [K], 1 atm", "TACOT B'c", "TACOT B'g", "CPh70 B'c", "CPh70 B'g", "écart sur B'c", ""])
for T, bt, bgt, bc, bgc in [(1000, 0.14221, 0.0388, 0.14774, 0.0205),
                            (2000, 0.16684, 0.0455, 0.17105, 0.0237),
                            (3000, 0.18962, 0.0517, 0.18564, 0.0257),
                            (3400, 0.24090, 0.0657, 0.22659, 0.0314),
                            (3600, 0.42054, 0.1147, 0.37210, 0.0515)]:
    r = row(ws, r, [T, bt, bgt, bc, bgc, f"=(E{r}-C{r})/C{r}", ""],
            fonts=[F_LAB, F_IN, F_IN, F_IN, F_IN, F_CALC, None],
            nfs=["0", "0.00000", "0.0000", "0.00000", "0.0000", NF_E])
r = note(ws, r, "Valeurs produites par `python cph70_bprime.py` et `python tacot_bprime/material_response.py`. "
                "Sous 3000 K l'écart reste sous ~4 % ; il explose au genou de sublimation, où la courbe "
                "B'c(B'g) devient raide. Moins de soufflage pyrolytique ⇒ le char du CPh70 survit à plus "
                "haute température.")
r += 1

r = section(ws, r, "§6 — SENSIBILITÉ AU RAPPORT FIBRES/RÉSINE")
rsn = r + 1
r = headrow(ws, r, ["f_fibres (volume solide)", "ρ_v [kg/m³]", "ρ_c [kg/m³]", "k = B'g/B'c",
                    "gaz de pyrolyse C", "gaz de pyrolyse H", "gaz de pyrolyse O"])
for ff in [0.10, 0.30, 0.50, 0.70, 0.90]:
    r = row(ws, r, [ff,
                    f"=(1-$C${rphi})*(B{r}*$C${rrhof}+(1-B{r})*$C${rrhom})",
                    f"=(1-$C${rphi})*(B{r}*$C${rrhof}+$C${ry}*(1-B{r})*$C${rrhom})",
                    f"=(C{r}-D{r})/D{r}", f"=$C${rgc}", f"=$D${rgc}", f"=$E${rgc}"],
            fonts=[F_IN, F_CALC, F_CALC, F_CALC, F_LINK, F_LINK, F_LINK],
            nfs=[NF_X, "0.0", "0.0", "0.0000", NF_X, NF_X, NF_X])
r = note(ws, r, "Les trois dernières colonnes sont CONSTANTES : c'est toute la démonstration de cet onglet. "
                "Le rapport fibres/résine déplace k d'un facteur 5 sur la plage balayée, et ne touche pas "
                "d'un iota la composition portée dans le XML.")
r += 1

r = section(ws, r, "§7 — RÉPONSE MATÉRIAU (n'entre PAS dans le XML)")
r = kv(ws, r, "k du CPh70", f"=C{rk}", "", font=F_CALC, nf="0.0000")
r = kv(ws, r, "k du TACOT (rappel)", f"='TACOT'!C{TAC_XML[1]}", "", font=F_LINK, nf="0.0000")
r = kv(ws, r, "rapport ρ_char CPh70 / ρ_char TACOT", f"=C{rrc}/'TACOT'!C{TAC_XML[3]}",
       "à B'c égal, la récession ṡ = B'c·ṁe/ρ_c du CPh70 est 5.85 fois plus lente", font=F_CALC, nf="0.00")
r += 1

r = section(ws, r, "§8 — BLOC XML")
r = xmlblock(ws, r, [
    '<mixture thermo_db="NASA-9" >',
    '    <species>',
    '       C H O N CH4 CN CO CO2 C2 C2H C2H2,acetylene C3 C4 C4H2,butadiyne C5',
    '       HCN H2 H2O N2 CH2OH CNN CNC CNCOCN C6H6 HNC C(gr)',
    '    </species>',
    '    <element_compositions default="air">',
    '        <composition name="air">N:0.79, O:0.21</composition>',
    '        <composition name="cph70_pyro">C:0.206, H:0.679, O:0.115</composition>',
    '        <composition name="cph70_char">C:1.0</composition>',
    '    </element_compositions>',
    '</mixture>',
])
CPH_XML = (rxml, rk)


# ===========================================================================
# ONGLET  ZURAM
# ===========================================================================
ws = ws_zur
setup(ws, "7030A0")
r = title(ws, "ZURAM® 18/50 — carbone/phénolique DLR",
          "Chaîne INVERSE du TACOT : la donnée primaire est une composition élémentaire MASSIQUE mesurée, pas une spéciation moléculaire",
          "data/mixtures/zuram-air.xml · zuram-pyrogas.xml")

r = section(ws, r, "§0 — FICHE D'IDENTITÉ")
r = kv(ws, r, "Matériau", "ZURAM® 18/50, ablateur carbone/phénolique du DLR", "'ZURAM_official'!A4 du classeur Tacot_Zuram_Calcarb_database_v4.3.1.ods")
r = kv(ws, r, "Renfort", "préforme Mersen Calcarb CBCF 18/2000 — « 18 » = 180 kg/m³ nominal", "'Calcarb_official'!D15 ; note de version 4.3.0")
r = kv(ws, r, "Résine", "phénolique catalysée à l'hexamine (HMTA) — novolac DÉDUIT, non déclaré", "[THo] Aerosp. Sci. Technol. 119 (2021) 107079, § 4.2")
r = kv(ws, r, "« 50 » du nom", "teneur en résine, en % masse de la recette — CONFIRMÉ", "Pagan et al. 2017, table 1 : les variantes −25 % et −50 % de résine s'appellent 18/43 et 18/33, soit exactement 42.857 % et 33.333 %")
r = kv(ws, r, "Source du gaz", "campagne AblaNTIS du VKI, fractions MASSIQUES élémentaires", "'ZURAM_official'!E87:E89")
r = kv(ws, r, "Compositions du XML", "BLedge (-bl) · VKIZuramPyroGas (-py) · Char (-char, -char-elem C)", "reproduit le cas de référence AblaNTIS carbonPhenolInAir")
r += 1

r = section(ws, r, "§1 — HYPOTHÈSES")
r = hyp(ws, r, "H1", "La donnée primaire est la composition élémentaire MASSIQUE du gaz, mesurée — il n'y a pas de spéciation moléculaire à sommer.",
        "C'est la différence de fond avec le TACOT : ici le calcul se réduit à une conversion masse → mole. La contrepartie est que la traçabilité s'arrête à trois décimales, la source [4] VKI_ZURAM_characterization n'étant pas publique.")
r = hyp(ws, r, "H2", "Conversion effectuée avec les masses molaires DU CLASSEUR (C 12.0107, H 1.00794, O 15.999).",
        "Elles seules reproduisent 'ZURAM_official'!D87:D89 au 4e-16 près. Les masses usuelles donneraient le même résultat à l'arrondi près, mais pas la reproduction exacte.")
r = hyp(ws, r, "H3", "Le gaz est déclaré sans azote, bien que la résine en contienne 1.44 % (hexamine).",
        "[THo] mesure un rendement massique constant de 1.5 % de NH3 au-dessus de 500 °C. Comme pour le TACOT, l'azote est traité comme un composé de catalyseur, hors squelette.")
r = hyp(ws, r, "H4", "Le char est du carbone pur : C:1.0.",
        "[THo] mesure pourtant C 96.36 / N 0.13 / H 0.08 / O 3.43 % à 800 °C, et le souligne explicitement. L'hypothèse reste licite en ablation (paroi > 2000 K) et c'est elle qui ferme le mieux le bilan (§5, contrôle C).")
r = hyp(ws, r, "H5", "Microstructure : densité intrinsèque de matrice INCHANGÉE à la pyrolyse (pure perte de volume).",
        "Anomalie ouverte n° 3 du classeur : « Intrinsic density of charred resin not well evaluated by He pycnometry ». N'affecte que ρ_c et donc k, pas le XML.")
r = hyp(ws, r, "H6", "Liste d'espèces = les 20 espèces gazeuses du cas AblaNTIS + C(gr).",
        "Le but est de reproduire EXACTEMENT la table B' de référence Bprime_carbonPhenolInAir_AblaNTIS.txt : on reprend sa liste à l'identique.")
r += 1

r = section(ws, r, "§2 — DONNÉES D'ENTRÉE")
r = step(ws, r, "Gaz de pyrolyse — fractions massiques élémentaires mesurées", "'ZURAM_official'!E87:E89")
rym = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "", ""])
r = row(ws, r, ["y_E — fraction massique [-]", 0.457, 0.162, 0.381, f"=SUM(C{r}:E{r})", "", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_CALC, None, None],
        nfs=[None, NF_X3, NF_X3, NF_X3, NF_X3], fills=[None, FILL_KEY, FILL_KEY, FILL_KEY, None])
r = row(ws, r, ["M_E [g/mol] — masses du classeur", f"={MCz}", f"={MHz}", f"={MOz}", "", "", ""],
        fonts=[F_LAB, F_LINK, F_LINK, F_LINK, None, None, None],
        nfs=[None, "0.00000", "0.00000", "0.00000"])
r = row(ws, r, ["x_E du classeur ('ZURAM_official'!D87:D89)",
                0.170941536278466, 0.722071254424386, 0.106987209297148, "", "", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, None, None, None],
        nfs=[None, "0.000000000000", "0.000000000000", "0.000000000000"])
rxref = r - 1
r += 1
r = step(ws, r, "Résine vierge et char — campagne [THo] (analyseur élémentaire, fig. 7)", "% masse")
rres = r + 1
r = headrow(ws, r, ["", "C", "H", "N", "O", "Somme", ""])
r = row(ws, r, ["résine vierge [% masse]", 75.21, 5.75, 1.44, 14.13, f"=SUM(C{r}:F{r})", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_IN, F_CALC, None],
        nfs=[None, NF_P, NF_P, NF_P, NF_P, NF_P])
r = row(ws, r, ["char à 800 °C [% masse]", 96.36, 0.08, 0.13, 3.43, f"=SUM(C{r}:F{r})", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_IN, F_CALC, None],
        nfs=[None, NF_P, NF_P, NF_P, NF_P, NF_P])
rchar800 = r - 1
r = note(ws, r, "La somme de l'analyse de résine vaut 96.5 % (relevé graphique de la fig. 7) : l'étape 1 "
                "du contrôle C la renormalise à 100 %. [THo] commente lui-même la discordance sur l'oxygène "
                "entre analyseur élémentaire et micropyrolyse, imputée aux composés lourds non détectés.")
r += 1
r = step(ws, r, "Microstructure — 'ZURAM_official'!D12:F21", "fractions volumiques et densités intrinsèques")
rvf = r + 1
r = headrow(ws, r, ["Fractions volumiques", "pores", "fibres", "résine", "Somme", "ρ intrinsèque [g/cm³]", ""])
r = row(ws, r, ["vierge", 0.707299, 0.129973, 0.162729, f"=SUM(C{r}:E{r})", "", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_CALC, None, None],
        nfs=[None, "0.000000", "0.000000", "0.000000", NF_X3])
r = row(ws, r, ["char", 0.769174, 0.129973, 0.100854, f"=SUM(C{r}:E{r})", "", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_CALC, None, None],
        nfs=[None, "0.000000", "0.000000", "0.000000", NF_X3])
r = row(ws, r, ["densité intrinsèque [g/cm³]", "—", 1.5772, 1.3151, "", "", ""],
        fonts=[F_LAB, F_NOTE, F_IN, F_IN, None, None, None],
        nfs=[None, None, "0.0000", "0.0000"])
rrho = r - 1
r = note(ws, r, "Piège d'unités : l'en-tête du classeur annonce [kg/m³] mais les valeurs sont en g/cm³. "
                "Anomalie ouverte n° 2 : 0.129973 × 1577.2 = 205.0 kg/m³ de fibres, soit 14 % de plus que "
                "les 180 kg/m³ nominaux de la préforme CBCF 18/2000.")
r += 1
r = step(ws, r, "Cinétique de pyrolyse — 'ZURAM_official'!B80:B83", "les f_i ont été divisées par la fraction massique de résine pour porter sur la résine seule (B77)")
rf = r + 1
r = headrow(ws, r, ["Réaction", "f_i [-]", "log10 A", "E/R [K]", "m", "", ""])
for i, (f, la, er, m) in enumerate([(0.035070, 5.33, 8178.5, 4.30),
                                    (0.027687, 8.69, 16068.4, 3.70),
                                    (0.095981, 10.60, 21612.9, 2.57),
                                    (0.221495, 11.67, 26423.8, 4.63)], 1):
    r = row(ws, r, [f"réaction {i}", f, la, er, m, "", ""],
            fonts=[F_LAB, F_IN, F_IN, F_IN, F_IN, None, None],
            nfs=[None, "0.000000", "0.00", "0.0", "0.00"])
rf_end = r - 1
r = row(ws, r, ["Σ f_i", f"=SUM(C{rf}:C{rf_end})", "", "", "", "", ""],
        fonts=[F_LABB, F_CALC, None, None, None, None, None], nfs=[None, "0.000000"])
rfsum = r - 1
r += 1

r = section(ws, r, "§3 — CALCUL PAS À PAS")
r = step(ws, r, "Étape 1 — conversion masse → mole (c'est TOUT le calcul du XML)",
         "n_E = y_E / M_E ;  x_E = n_E / Σ_j n_j")
rn = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "M_gaz [g/mol]", "écart / classeur"])
r = row(ws, r, ["n_E = y_E / M_E  [mol par gramme de gaz]",
                f"=C{rym}/C{rym+1}", f"=D{rym}/D{rym+1}", f"=E{rym}/E{rym+1}",
                f"=SUM(C{r}:E{r})", f"=1/F{r}", ""],
        nfs=[None, "0.00000", "0.00000", "0.00000", "0.00000", NF_M])
rnn = r - 1
r = row(ws, r, ["x_E = n_E / Σ n_j", f"=C{rnn}/$F${rnn}", f"=D{rnn}/$F${rnn}", f"=E{rnn}/$F${rnn}",
                f"=SUM(C{r}:E{r})", "",
                f"=MAX(ABS(C{r}-C{rxref}),ABS(D{r}-D{rxref}),ABS(E{r}-E{rxref}))"],
        nfs=[None, "0.000000000000", "0.000000000000", "0.000000000000", NF_X3, None, "0.00E+00"])
rxE = r - 1
r = note(ws, r, "M_gaz = 1 / Σ(y_E/M_E) = 4.4926 g/mol, contre 4.9990 pour le TACOT : le gaz du ZURAM est "
                "plus léger — plus riche en hydrogène (0.722 contre 0.679) et nettement plus pauvre en "
                "carbone (0.171 contre 0.206). L'écart avec le classeur (colonne H) vaut ~4e-16.")
r += 1

r = section(ws, r, "§4 — RÉSULTAT : CE QUI ENTRE DANS LE XML")
r, rcalc, rxml = result_block(ws, r, "VKIZuramPyroGas", f"C{rxE}", f"D{rxE}", f"E{rxE}",
                              (0.171, 0.722, 0.107))
r = note(ws, r, "L'écart est celui de l'arrondi à trois décimales. ATTENTION : l'onglet ZURAM_preliminary "
                "du même classeur porte une valeur DIFFÉRENTE (C:0.1246, H:0.7506, O:0.1249), issue d'une "
                "autre voie — 37 % d'écart sur le carbone. Ce n'est PAS la valeur du XML.")
r += 1
r = headrow(ws, r, ["Char (Char)", "C", "H", "O", "", "", ""])
r = row(ws, r, ["composition du XML", 1.0, "—", "—", "", "", ""],
        fonts=[F_LAB, F_OUT, F_NOTE, F_NOTE, None, None, None],
        nfs=[None, NF_X3, None, None], fills=[None, FILL_OUT, None, None])
r += 1

r = section(ws, r, "§5 — CONTRÔLES ET VALIDATION CROISÉE")
r = step(ws, r, "Contrôle A — le rendement en char par DEUX voies indépendantes du classeur",
         "(a) volumique : ε_m,char/ε_m,vierge   (b) cinétique : 1 − Σ f_i")
r = kv(ws, r, "(a) fractions volumiques de matrice", f"=E{rvf+1}/E{rvf}",
       "0.100854 / 0.162729 — licite car la densité intrinsèque de matrice est supposée inchangée (H5)",
       font=F_CALC, nf="0.000000")
rya = r - 1
r = kv(ws, r, "(b) cinétique de pyrolyse", f"=1-C{rfsum}", "1 − Σ f_i sur les 4 réactions",
       font=F_CALC, nf="0.000000")
ryb = r - 1
r = kv(ws, r, "écart entre les deux voies", f"=ABS(C{rya}-C{ryb})", "attendu ~4e-16 — le classeur est cohérent à la dernière décimale",
       font=F_CALC, nf="0.00E+00")
r = note(ws, r, "Rendement en char de la résine du ZURAM : 0.6198, contre 0.50 pour le TACOT. Le ZURAM "
                "charbonne nettement plus — cohérent avec une résine plus réticulée (motif C7H6O plutôt "
                "que C6H6O). Comme pour le TACOT, aucune cellule ne porte ce nombre : il est implicite.")
r += 1
r = step(ws, r, "Contrôle C — fermeture croisée résine − char → gaz (deux campagnes distinctes)",
         "n_E(gaz) = n_E(résine renormalisée) − n_E(char)")
rcl = r + 1
r = headrow(ws, r, ["Base 100 g de résine vierge", "C", "H", "O", "Somme", "écart max / XML", ""])
r = row(ws, r, ["résine renormalisée à 100 % [% masse]",
                f"=C{rres}/$G${rres}*100", f"=D{rres}/$G${rres}*100", f"=F{rres}/$G${rres}*100", "", "", ""],
        nfs=[None, NF_P, NF_P, NF_P])
rrn = r - 1
r = row(ws, r, ["moles d'atomes de la résine", f"=C{rrn}/{MC}", f"=D{rrn}/{MH}", f"=E{rrn}/{MO}",
                f"=SUM(C{r}:E{r})", "", ""], nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rmr = r - 1
r = row(ws, r, ["char retranché — carbone pur, Y = 0.6198",
                f"=$C${rya}*100/{MC}", 0, 0, f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rc1 = r - 1
r = row(ws, r, ["gaz obtenu — x_E",
                f"=(C{rmr}-C{rc1})/($F${rmr}-$F${rc1})", f"=(D{rmr}-D{rc1})/($F${rmr}-$F${rc1})",
                f"=(E{rmr}-E{rc1})/($F${rmr}-$F${rc1})", f"=SUM(C{r}:E{r})",
                f"=MAX(ABS(C{r}-$C${rxml})/$C${rxml},ABS(D{r}-$D${rxml})/$D${rxml},ABS(E{r}-$E${rxml})/$E${rxml})", ""],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3, NF_E])
r = row(ws, r, ["char retranché — char mesuré à 800 °C",
                f"=$C${rya}*100*C{rchar800}/100/{MC}", f"=$C${rya}*100*D{rchar800}/100/{MH}",
                f"=$C${rya}*100*F{rchar800}/100/{MO}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rc2 = r - 1
r = row(ws, r, ["gaz obtenu — x_E",
                f"=(C{rmr}-C{rc2})/($F${rmr}-$F${rc2})", f"=(D{rmr}-D{rc2})/($F${rmr}-$F${rc2})",
                f"=(E{rmr}-E{rc2})/($F${rmr}-$F${rc2})", f"=SUM(C{r}:E{r})",
                f"=MAX(ABS(C{r}-$C${rxml})/$C${rxml},ABS(D{r}-$D${rxml})/$D${rxml},ABS(E{r}-$E${rxml})/$E${rxml})", ""],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3, NF_E])
r = note(ws, r, "Accord à 4.9 % avec l'hypothèse de char de carbone pur, contre 10.4 % avec le char "
                "réellement mesuré. La composition du gaz vient d'une source VKI, celle de la résine et du "
                "char de la campagne [THo] : la fermeture est un VRAI contrôle, pas une tautologie. "
                "Le fait que le char pur ferme le mieux suggère que la donnée VKI a elle aussi été établie "
                "sous cette hypothèse.")
r += 1
r = step(ws, r, "Contrôle D — le motif équivalent de la résine", "renormalisation hors azote et comparaison aux motifs de l'onglet Constantes")
rmot = r + 1
r = headrow(ws, r, ["Composition hors azote [% masse]", "C", "H", "O", "écart max au motif", "", ""])
r = row(ws, r, ["ZURAM mesuré, renormalisé hors N",
                f"=C{rres}/(C{rres}+D{rres}+F{rres})*100", f"=D{rres}/(C{rres}+D{rres}+F{rres})*100",
                f"=F{rres}/(C{rres}+D{rres}+F{rres})*100", "", "", ""],
        nfs=[None, NF_P, NF_P, NF_P])
rzm = r - 1
for lab, rr in [("phénol C6H6O", 25), ("novolac linéaire C7H6O", 26),
                ("résol saturé C7.5H6O", 27)]:
    r = row(ws, r, [lab, f"=Constantes!$F${rr}", f"=Constantes!$G${rr}", f"=Constantes!$H${rr}",
                    f"=MAX(ABS(C{r}-$C${rzm}),ABS(D{r}-$D${rzm}),ABS(E{r}-$E${rzm}))", "", ""],
            fonts=[F_LAB, F_LINK, F_LINK, F_LINK, F_CALC, None, None],
            nfs=[None, NF_P, NF_P, NF_P, NF_P])
r = note(ws, r, "Le ZURAM tombe à 0.35 point du novolac linéaire C7H6O — contre 0.85 point du phénol C6H6O "
                "pour le TACOT. Les deux résines emploient très vraisemblablement la même chimie ; c'est la "
                "quantité de composés piégés (eau, HMTA) qui distingue les deux analyses. Combiné à "
                "l'hexamine, c'est ce qui fait déduire « novolac ».")
r += 1

r = section(ws, r, "§6 — SENSIBILITÉ")
r = para(ws, r, "Le gaz du XML ne dépend d'AUCUNE des hypothèses de rendement ou de motif : il est mesuré "
                "directement (§3, une seule étape). Les sensibilités portent donc sur les CONTRÔLES, pas "
                "sur le résultat — c'est l'inverse du SC-1008 et du liège, où le rendement en char entre "
                "dans la chaîne de calcul du XML.", 30)
rsn = r + 1
r = headrow(ws, r, ["Rendement en char testé", "C", "H", "O", "écart max / XML", "", ""])
for Y in [0.55, 0.60, 0.619766, 0.65]:
    r = row(ws, r, [Y,
                    f"=($C${rmr}-B{r}*100/{MC})/($F${rmr}-B{r}*100/{MC})",
                    f"=$D${rmr}/($F${rmr}-B{r}*100/{MC})",
                    f"=$E${rmr}/($F${rmr}-B{r}*100/{MC})",
                    f"=MAX(ABS(C{r}-$C${rxml})/$C${rxml},ABS(D{r}-$D${rxml})/$D${rxml},ABS(E{r}-$E${rxml})/$E${rxml})", "", ""],
            fonts=[F_IN, F_CALC, F_CALC, F_CALC, F_CALC, None, None],
            nfs=["0.000000", NF_X, NF_X, NF_X, NF_E])
r = note(ws, r, "Le minimum d'écart tombe près de la valeur du classeur — le rendement en char déduit de la "
                "microstructure est compatible avec la composition du gaz mesurée indépendamment.")
r += 1

r = section(ws, r, "§7 — RÉPONSE MATÉRIAU (n'entre PAS dans le XML)")
r = kv(ws, r, "ρ vierge [kg/m³]", f"=(D{rvf}*D{rrho}+E{rvf}*E{rrho})*1000",
       "'ZURAM_official'!D17 — attendu 419.0", font=F_CALC, nf="0.0")
rrv = r - 1
r = kv(ws, r, "ρ char [kg/m³]", f"=(D{rvf+1}*D{rrho}+E{rvf+1}*E{rrho})*1000",
       "'ZURAM_official'!E17 — attendu 337.6", font=F_CALC, nf="0.0")
rrc = r - 1
r = kv(ws, r, "fibres seules [kg/m³]", f"=D{rvf}*D{rrho}*1000",
       "205.0 au lieu des 180 nominaux — anomalie ouverte n° 2 du classeur", font=F_CALC, nf="0.0")
r = kv(ws, r, "résine en % de la masse vierge", f"=E{rvf}*E{rrho}*1000/C{rrv}*100",
       "51.07 % (TACOT : 42.9 %) pour 16.3 % du volume", font=F_CALC, nf=NF_P)
r = kv(ws, r, "gaz libéré [kg/m³]", f"=C{rrv}-C{rrc}", "81.4 kg/m³, soit 19.4 % de la masse vierge (TACOT : 21.4 %)",
       font=F_CALC, nf="0.0")
r = kv(ws, r, "couplage k = B'g / B'c", f"=(C{rrv}-C{rrc})/C{rrc}", "attendu 0.2410 (TACOT : 0.2727)",
       font=F_CALC, nf="0.0000")
rk = r - 1
r += 1

r = section(ws, r, "§8 — BLOC XML")
r = xmlblock(ws, r, [
    '<mixture thermo_db="NASA-9">',
    '    <species>',
    '       C(gr) C H O N CH CH4 CO CO2 CN C2H C2H2,acetylene C3 C3H C4',
    '       C4H2,butadiyne C5 HCN H2 H2O N2',
    '    </species>',
    '    <element_compositions default="BLedge">',
    '        <composition name="BLedge">           O:0.21, N:0.79            </composition>',
    '        <composition name="VKIZuramPyroGas">  C:0.171, H:0.722, O:0.107  </composition>',
    '        <composition name="Char">             C:1.0                     </composition>',
    '    </element_compositions>',
    '</mixture>',
])
ZUR_XML = (rxml, rk)


# ===========================================================================
# ONGLET  SC-1008
# ===========================================================================
ws = ws_sc
setup(ws, "00B050")
r = title(ws, "SC-1008 — résol phénol-formaldéhyde (matrice du PICA)",
          "Gaz obtenu par FERMETURE ÉLÉMENTAIRE sur un motif chimique idéalisé, validé contre une mesure indépendante",
          "data/mixtures/sc1008-air.xml · sc1008-pyrogas.xml")

r = section(ws, r, "§0 — FICHE D'IDENTITÉ")
r = kv(ws, r, "Résine", "Durite™ SC-1008, Hexion / Bakelite Synthetics (ex-Borden)", "[Wo] J. Anal. Appl. Pyrolysis 122 (2016) 258-267, § 2.2")
r = kv(ws, r, "Type", "RÉSOL phénol-formaldéhyde, monocomposant, réticulant par chauffage seul", "excès de formaldéhyde, catalyse basique : porte déjà ses méthylols, durcit sans durcisseur")
r = kv(ws, r, "Emploi", "matrice du PICA, sur préforme carbone FiberForm", "densité PICA 0.274 g/cm³ ± 10 %, porosité 0.8 ± 10 %")
r = kv(ws, r, "Livraison", "20-25 % isopropanol · 11-18 % phénol libre · 0.6-2 % formaldéhyde libre", "[Wo] § 2.2 — le phénol libre part en région 1 sans venir du réseau")
r = kv(ws, r, "Sources", "[TS] Trick & Saliba, Carbon 33 (1995) 1509-1515 · [Wo] Wong et al. 2016", "sc1008_bprime/resine_sc1008.md")
r = kv(ws, r, "Compositions du XML", "air (-bl) · sc1008_pyro (-py) · sc1008_char (-char, -char-elem C)", "")
r += 1

r = section(ws, r, "§1 — HYPOTHÈSES")
r = hyp(ws, r, "H1", "Une résine résol cuite = un cycle phénolique + ses ponts méthylène : motif C(6+b)H6O.",
        "Le nombre de ponts b n'est pas libre : le phénol a 3 sites réactifs (2 ortho + 1 para) et un pont est partagé entre 2 cycles, d'où b = 1.5 à réseau saturé.")
r = hyp(ws, r, "H2", "PAS D'AZOTE dans le gaz — et cette fois ce n'est pas une approximation.",
        "Un résol réticule sans durcisseur : il n'y a pas d'hexaméthylènetétramine dans la formulation. Le gaz est NATIVEMENT C/H/O. L'argument que le TACOT (2.35 % N) et le ZURAM (1.4 %) doivent construire est ici sans objet.")
r = hyp(ws, r, "H3", "Le char est du carbone pur ⇒ H et O traversent la pyrolyse sans être touchés.",
        "Conséquence remarquable : H/O(gaz) = H/O(résine) INDÉPENDAMMENT du rendement en char. C'est l'invariant du §5, un test qui ne dépend d'aucune hypothèse fragile.")
r = hyp(ws, r, "H4", "Rendement en char Y = 0.55.",
        "Ce n'est PAS la moyenne des trois sources (0.545 / 0.55-0.60 / 0.62) : le milieu arithmétique 0.575 s'écarte de [TS] de 8.9 %. 0.55 est la seule valeur qui tombe dans la fourchette ATG ET reproduise la composition de [TS] à 1.5 %.")
r = hyp(ws, r, "H5", "La borne haute Y = 0.62 de [Wo] est PROPRE AU PICA et ne vaut pas pour la résine seule.",
        "La préforme carbone capte des radicaux pendant la pyrolyse et relève le rendement — [Wo] l'observe directement : H2 divisé par ~5 par rapport à la résole nue.")
r = hyp(ws, r, "H6", "Liste d'espèces = les 25 de la table B' de référence + phénol + C(gr).",
        "Le phénol et le crésol sont les produits dominants EN MASSE de la pyrolyse d'un résol ([TS] région 1 ; [Wo] table 1) : les omettre appauvrirait l'équilibre pariétal.")
r += 1

r = section(ws, r, "§2 — DONNÉES D'ENTRÉE")
r = kv(ws, r, "nombre de ponts méthylène par cycle, b", 1.5, "3 sites réactifs / 2 cycles par pont — cf. H1", key=True)
rb = r - 1
r = kv(ws, r, "rendement en char de la résine, Y", 0.55, "recoupement [TS] (0.545) et ATG SC-1008 nue (0.55-0.60) — cf. H4", key=True)
rY = r - 1
r = kv(ws, r, "composition mesurée de [TS] — C", 0.2564, "reconstruite au §5 depuis les tables 2 et 4 de Trick & Saliba (LMS = 50/50 phénol/crésol)")
rTS = r - 1
r = kv(ws, r, "                                H", 0.6388, "")
r = kv(ws, r, "                                O", 0.1048, "")
r += 1

r = section(ws, r, "§3 — CALCUL PAS À PAS")
r = step(ws, r, "Étape 1 — construire le motif de la résine cuite",
         "phénol C6H6O + b·CH2O − b·H2O")
rmo = r + 1
r = headrow(ws, r, ["Brique", "coefficient", "ν_C", "ν_H", "ν_O", "", ""])
r = row(ws, r, ["phénol C6H6O", 1, 6, 6, 1, "", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_IN, None, None], nfs=[None, NF_X3, "0", "0", "0"])
r = row(ws, r, ["+ formaldéhyde CH2O", f"=$C${rb}", 1, 2, 1, "", ""],
        fonts=[F_LAB, F_CALC, F_IN, F_IN, F_IN, None, None], nfs=[None, NF_X3, "0", "0", "0"])
r = row(ws, r, ["− eau de condensation H2O", f"=-$C${rb}", 0, 2, 1, "", ""],
        fonts=[F_LAB, F_CALC, F_IN, F_IN, F_IN, None, None], nfs=[None, NF_X3, "0", "0", "0"])
rmo_end = r - 1
r = row(ws, r, ["MOTIF CUIT — ν_E = Σ coeff · ν",
                "", f"=SUMPRODUCT($C${rmo}:$C${rmo_end},D{rmo}:D{rmo_end})",
                f"=SUMPRODUCT($C${rmo}:$C${rmo_end},E{rmo}:E{rmo_end})",
                f"=SUMPRODUCT($C${rmo}:$C${rmo_end},F{rmo}:F{rmo_end})", "", ""],
        fonts=[F_LABB, None, F_CALC, F_CALC, F_CALC, None, None],
        nfs=[None, None, NF_X3, NF_X3, NF_X3])
rmot = r - 1
r = note(ws, r, "Avec b = 1.5 : C7.5H6O. Noter que H et O sont RIGOUREUSEMENT insensibles à b "
                "(+2b des formaldéhydes, −2b des eaux) : seul le carbone bouge. C'est la raison profonde "
                "de l'invariant H/O du §5.")
r = step(ws, r, "Étape 2 — masse molaire et composition massique du motif",
         "M = ν_C·M_C + ν_H·M_H + ν_O·M_O ;  w_E = ν_E·M_E / M")
rM = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "M [g/mol]", "", ""])
r = row(ws, r, ["ν_E · M_E [g/mol]", f"=D{rmot}*{MC}", f"=E{rmot}*{MH}", f"=F{rmot}*{MO}",
                f"=SUM(C{r}:E{r})", "", ""], nfs=[None, NF_M, NF_M, NF_M, NF_M])
rMr = r - 1
r = row(ws, r, ["w_E [% masse]", f"=C{rMr}/$F${rMr}*100", f"=D{rMr}/$F${rMr}*100",
                f"=E{rMr}/$F${rMr}*100", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_P, NF_P, NF_P, NF_P])
r = note(ws, r, "Attendu : M = 112.13 g/mol, soit C 80.3 % / H 5.4 % / O 14.3 % en masse. "
                "L'écart résol (b = 1.5) / novolac (b = 1) ne vaut que 1.1 point sur C — plus PETIT que "
                "la dispersion entre deux novolacs réellement mesurés (TACOT 77.4 vs ZURAM 79.1). "
                "L'analyse C/H/O ne permet donc PAS d'identifier le type de résine ; seul l'azote le fait.")
r = step(ws, r, "Étape 3 — fermeture élémentaire sur 100 g de résine : gaz = résine − char",
         "n_E(gaz) = n_E(résine) − n_E(char) ;  n_C(char) = Y·100 / M_C")
rfe = r + 1
r = headrow(ws, r, ["Base 100 g de résine vierge", "C", "H", "O", "Somme", "masse [g]", ""])
r = row(ws, r, ["résine — moles d'atomes",
                f"=100/$F${rMr}*D{rmot}", f"=100/$F${rMr}*E{rmot}", f"=100/$F${rMr}*F{rmot}",
                f"=SUM(C{r}:E{r})", f"=C{r}*{MC}+D{r}*{MH}+E{r}*{MO}", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL, "0.00"])
rre = r - 1
r = row(ws, r, ["char — carbone pur, Y = 0.55", f"=$C${rY}*100/{MC}", 0, 0,
                f"=SUM(C{r}:E{r})", f"=$C${rY}*100", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL, "0.00"])
rch = r - 1
r = row(ws, r, ["gaz — moles d'atomes", f"=C{rre}-C{rch}", f"=D{rre}-D{rch}", f"=E{rre}-E{rch}",
                f"=SUM(C{r}:E{r})", f"=F{rre}-F{rch}", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL, "0.00"])
rga = r - 1
r = step(ws, r, "Étape 4 — normaliser en fractions molaires, et convertir en fractions massiques",
         "x_E = n_E / Σ n_j ;  y_E = x_E·M_E / Σ x_j·M_j")
rxE_h = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "M_gaz [g/mol]", "H/O"])
r = row(ws, r, ["x_E (fraction molaire élémentaire)", f"=C{rga}/$F${rga}", f"=D{rga}/$F${rga}",
                f"=E{rga}/$F${rga}", f"=SUM(C{r}:E{r})",
                f"=C{r}*{MC}+D{r}*{MH}+E{r}*{MO}", f"=D{r}/E{r}"],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3, NF_M, "0.000"])
rxE = r - 1
r = row(ws, r, ["y_E (fraction massique élémentaire)", f"=C{rxE}*{MC}/$G${rxE}", f"=D{rxE}*{MH}/$G${rxE}",
                f"=E{rxE}*{MO}/$G${rxE}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_X3, NF_X3, NF_X3, NF_X3])
r = note(ws, r, "La ligne massique est l'équivalent strict à déclarer si l'on écrit "
                "<composition name=\"sc1008_pyro\" type=\"mass\">C:0.563, H:0.120, O:0.317</composition>. "
                "Sans l'attribut type=\"mass\", le parseur suppose du molaire — et le mélange serait faux.")
r += 1

r = section(ws, r, "§4 — RÉSULTAT : CE QUI ENTRE DANS LE XML")
r, rcalc, rxml = result_block(ws, r, "sc1008_pyro", f"C{rxE}", f"D{rxE}", f"E{rxE}",
                              (0.2526, 0.6407, 0.1068))
r = headrow(ws, r, ["Char (sc1008_char)", "C", "H", "O", "", "", ""])
r = row(ws, r, ["composition du XML", 1.0, "—", "—", "", "", ""],
        fonts=[F_LAB, F_OUT, F_NOTE, F_NOTE, None, None, None],
        nfs=[None, NF_X3, None, None], fills=[None, FILL_OUT, None, None])
r += 1

r = section(ws, r, "§5 — CONTRÔLES ET VALIDATION CROISÉE")
r = step(ws, r, "Contrôle A — l'invariant H/O : il ne dépend PAS du rendement en char",
         "H/O(gaz) = H/O(résine), car un char de carbone pur ne prélève que du C")
rinv = r + 1
r = headrow(ws, r, ["Y testé", "C", "H", "O", "H/O", "verdict", ""])
for Y in [0.30, 0.50, 0.55, 0.70]:
    r = row(ws, r, [Y,
                    f"=($C${rre}-B{r}*100/{MC})/($F${rre}-B{r}*100/{MC})",
                    f"=$D${rre}/($F${rre}-B{r}*100/{MC})",
                    f"=$E${rre}/($F${rre}-B{r}*100/{MC})",
                    f"=D{r}/E{r}", "constant", ""],
            fonts=[F_IN, F_CALC, F_CALC, F_CALC, F_CALC, F_NOTE, None],
            nfs=[NF_X3, NF_X, NF_X, NF_X, "0.000", None])
r = note(ws, r, "Bornes : aucun motif phénolique concevable ne descend sous H/O = 3.6 (résol prépolymère "
                "non cuit C7.5H9O2.5), et ajouter de l'eau ne sauve rien — C7.5H6O·(H2O)w donne "
                "H/O = (6+2w)/(1+w), qui tend vers 2.0 PAR LE HAUT. C'est ce test qui a fait supprimer du "
                "dépôt le jeu PICA hérité de Goldstein, à H/O = 1.93 : il n'était reproductible par AUCUNE "
                "résine phénolique, à AUCUN degré de cuisson, avec AUCUNE quantité d'eau ajoutée. "
                "Valeurs du dépôt : tacot_pyro 5.90 · cph70_pyro 5.90 · VKIZuramPyroGas 6.75 · "
                "sc1008_pyro 6.00 · [TS] mesuré 6.10 · cork_pyro 4.89 (hors test : le liège n'est pas un phénolique).")
r += 1
r = step(ws, r, "Contrôle B — reconstruction INDÉPENDANTE de la mesure de [TS] : ni motif, ni rendement en char n'y entrent",
         "les tables 2 et 4 de [TS] se recoupent : f_ij·N_i / (f_kj·N_k) = d_ij / d_kj")
rt2 = r + 1
r = headrow(ws, r, ["Espèce", "table 2 — R1 [%]", "table 2 — R2 [%]", "table 2 — R3 [%]",
                    "table 4 — R1 [%]", "table 4 — R2 [%]", "table 4 — R3 [%]"])
TS = [("H2",      0.0, 59.4, 85.7,   0.0, 57.1, 42.9),
      ("H2O",    49.8, 12.7,  4.7,  66.3, 28.2,  5.5),
      ("LMS (phénol + crésol)", 50.1, 0.0, 0.0, 100.0, 0.0, 0.0),
      ("CO",      0.0, 12.7,  9.5,   0.0, 71.9, 28.1),
      ("CH4",     0.0, 14.9,  0.0,   0.0, 100.0, 0.0),
      ("CO2",     0.1,  0.2,  0.1,  20.5, 66.8, 12.7),
      ("C2H6",    0.0,  0.1,  0.0,   0.0, 100.0, 0.0)]
for lab, a, b_, c, d, e, f in TS:
    r = row(ws, r, [lab, a, b_, c, d, e, f],
            fonts=[F_LAB, F_IN, F_IN, F_IN, F_IN, F_IN, F_IN],
            nfs=[None, NF_P, NF_P, NF_P, NF_P, NF_P, NF_P])
rt2_end = r - 1
r = note(ws, r, "Table 2 = composition molaire DANS chaque région réactionnelle ; table 4 = distribution "
                "de chaque espèce SUR les régions (% du total). Régions : R1 300-550 °C, R2 400-800 °C, "
                "R3 560-900 °C. Aucune des deux ne donne la composition globale — mais elles se recoupent.")
rest = r + 1
r = headrow(ws, r, ["Estimation de N2/N3 par espèce commune", "valeur", "retenue ?", "", "", "", ""])
for lab, off in [("via H2O", 1), ("via H2", 0), ("via CO", 3), ("via CO2", 5)]:
    keep = "écartée : CO2 pèse 0.1-0.2 % dans la table 2, l'arrondi y domine le signal" if lab == "via CO2" else "oui"
    r = row(ws, r, [lab, f"=(G{rt2+off}/H{rt2+off})*(E{rt2+off}/D{rt2+off})", keep, "", "", "", ""],
            fonts=[F_LAB, F_CALC, F_NOTE, None, None, None, None], nfs=[None, "0.0000", None])
r = row(ws, r, ["moyenne des trois retenues", f"=AVERAGE(C{rest}:C{rest+2})", "accord à 1.2 %", "", "", "", ""],
        fonts=[F_LABB, F_CALC, F_NOTE, None, None, None, None], nfs=[None, "0.0000", None])
rN23 = r - 1
r = row(ws, r, ["N1/N2 via H2O (seule espèce commune à R1 et R2)",
                f"=(F{rt2+1}/G{rt2+1})*(D{rt2+1}/C{rt2+1})", "", "", "", "", ""],
        fonts=[F_LAB, F_CALC, None, None, None, None, None], nfs=[None, "0.0000"])
rN12 = r - 1
rNv = r + 1
r = headrow(ws, r, ["Tailles relatives des régions (N3 = 1)", "N1", "N2", "N3", "", "", ""])
r = row(ws, r, ["N_i", f"=C{rN12}*C{rN23}", f"=C{rN23}", 1, "", "", ""],
        fonts=[F_LAB, F_CALC, F_CALC, F_IN, None, None, None],
        nfs=[None, "0.0000", "0.0000", "0.0000"])
rN = r - 1
rgl = r + 1
r = headrow(ws, r, ["Composition moléculaire globale reconstruite", "n_j = Σ_i N_i·f_ij", "x_j",
                    "ν_C", "ν_H", "ν_O", ""])
NU = [(0, 2, 0), (0, 2, 1), (6.5, 7, 1), (1, 0, 1), (1, 4, 0), (1, 0, 2), (2, 6, 0)]
for i, (lab, *_rest) in enumerate(TS):
    nc, nh, no = NU[i]
    r = row(ws, r, [lab, f"=SUMPRODUCT($C${rN}:$E${rN},C{rt2+i}:E{rt2+i})",
                    f"=C{r}/$C${rgl+len(TS)}", nc, nh, no, ""],
            fonts=[F_LAB, F_CALC, F_CALC, F_IN, F_IN, F_IN, None],
            nfs=[None, "0.000", NF_X, NF_X3, NF_X3, NF_X3])
rgl_end = r - 1
r = row(ws, r, ["SOMME", f"=SUM(C{rgl}:C{rgl_end})", f"=SUM(D{rgl}:D{rgl_end})", "", "", "", ""],
        fonts=[F_LABB, F_CALC, F_CALC, None, None, None, None], nfs=[None, "0.000", NF_X3])
r = note(ws, r, "ν du LMS : phénol C6H6O et crésol C7H8O ne sont pas résolus par [TS]. "
                "L'hypothèse 50/50 donne C6.5H7O — c'est LA seule hypothèse vraiment libre de ce contrôle. "
                "Sensibilité : 100 % phénol → H/O 5.77 · 50/50 → 6.10 · 100 % crésol → 6.42. "
                "L'invariant H/O ≈ 6 tient sur toute la plage.")
rTSx = r + 1
r = headrow(ws, r, ["[TS] reconstruit — composition élémentaire", "C", "H", "O", "Somme", "H/O", "écart / XML"])
r = row(ws, r, ["n_E = Σ_j ν_E,j · x_j",
                f"=SUMPRODUCT($D${rgl}:$D${rgl_end},E{rgl}:E{rgl_end})",
                f"=SUMPRODUCT($D${rgl}:$D${rgl_end},F{rgl}:F{rgl_end})",
                f"=SUMPRODUCT($D${rgl}:$D${rgl_end},G{rgl}:G{rgl_end})",
                f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rTSn = r - 1
r = row(ws, r, ["x_E", f"=C{rTSn}/$F${rTSn}", f"=D{rTSn}/$F${rTSn}", f"=E{rTSn}/$F${rTSn}",
                f"=SUM(C{r}:E{r})", f"=D{r}/E{r}",
                f"=MAX(ABS(C{r}-$C${rxml})/$C${rxml},ABS(D{r}-$D${rxml})/$D${rxml},ABS(E{r}-$E${rxml})/$E${rxml})"],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3, "0.000", NF_E])
r = note(ws, r, "DEUX CHAÎNES DE RAISONNEMENT ENTIÈREMENT DISJOINTES — l'une partant d'une mesure de gaz "
                "déconvoluée en régions réactionnelles, l'autre d'un motif chimique idéalisé et d'un bilan "
                "de masse — convergent à 2 % : 1.5 % sur C, 0.3 % sur H, 1.9 % sur O. C'est la validation "
                "centrale de cette mise en données.")
r += 1
r = step(ws, r, "Contrôle C — pourquoi [Wo] n'est PAS utilisé comme composition élémentaire",
         "pourtant la mesure la plus directe : SC-1008 réel, 16 espèces, 20 paliers")
rwo = r + 1
r = headrow(ws, r, ["Traitement de la donnée [Wo]", "C", "H", "O", "H/O", "masse [mg/100 mg]", ""])
for lab, c, h, o, m in [("brut, table 1 intégrée", 0.120, 0.642, 0.238, 24.5),
                        ("A : − eau à T ≤ 664 K", 0.138, 0.639, 0.223, 21.2),
                        ("B : + eau > 850 K (attribution de [Wo])", 0.222, 0.622, 0.156, 13.0),
                        ("C : calage sur la balance", 0.154, 0.635, 0.211, 19.0)]:
    r = row(ws, r, [lab, c, h, o, f"=D{r}/E{r}", m, ""],
            fonts=[F_LAB, F_IN, F_IN, F_IN, F_CALC, F_IN, None],
            nfs=[None, NF_X3, NF_X3, NF_X3, "0.00", "0.0"])
r = note(ws, r, "Aucun traitement ne remonte à H/O = 6 : la fermeture de [Wo] n'est pas assurée. Trois causes, "
                "toutes documentées par les auteurs — (a) PICA est fortement hygroscopique et le GC totalise "
                "24.5 mg contre 19 mg à la balance, l'excédent valant exactement 0.306 mmol d'eau ; "
                "(b) le GC est limité à M < 400 g/mol et 35 % du carbone échappe à la mesure (les aromatiques "
                "ne pèsent que 3 % des moles mais 55 % du carbone mesuré) ; (c) composite ≠ résine. "
                "[Wo] reste la meilleure source pour la SPÉCIATION et le RENDEMENT EN CHAR, pas pour "
                "la composition élémentaire.")
r += 1

r = section(ws, r, "§6 — SENSIBILITÉ AU RENDEMENT EN CHAR")
rs = r + 1
r = headrow(ws, r, ["Y testé", "C", "H", "O", "écart sur C / [TS]", "source de la valeur", ""])
for Y, org in [(0.500, "trop bas — aucune source"),
               (0.545, "[TS], via la composition du gaz reconstruite"),
               (0.550, "RETENU — fourchette ATG ET accord [TS]"),
               (0.575, "milieu arithmétique des trois sources"),
               (0.600, "borne haute ATG"),
               (0.620, "[Wo] sur PICA — propre au composite, cf. H5")]:
    r = row(ws, r, [Y,
                    f"=($C${rre}-B{r}*100/{MC})/($F${rre}-B{r}*100/{MC})",
                    f"=$D${rre}/($F${rre}-B{r}*100/{MC})",
                    f"=$E${rre}/($F${rre}-B{r}*100/{MC})",
                    f"=ABS(C{r}-$C${rTS})/$C${rTS}", org, ""],
            fonts=[F_IN, F_CALC, F_CALC, F_CALC, F_CALC, F_NOTE, None],
            nfs=[NF_X3, NF_X, NF_X, NF_X, NF_E, None])
r = note(ws, r, "Le rendement en char est ici DANS la chaîne de calcul du XML (contrairement au TACOT et au "
                "ZURAM, où le gaz est mesuré directement) : ±0.05 sur Y déplace le carbone du gaz de ~12 %. "
                "C'est le paramètre à mesurer en priorité pour cette résine.")
r += 1

r = section(ws, r, "§7 — RÉPONSE MATÉRIAU (n'entre PAS dans le XML)")
r = kv(ws, r, "densité du PICA [g/cm³]", 0.274, "[Wo] § 2.2, ± 10 % ; porosité 0.8 ± 10 %")
r = kv(ws, r, "fraction massique de résine dans le PICA", 0.50, "[Wo] § 3.2 — la préforme carbone ne perd pas de masse", key=True)
rwr = r - 1
r = kv(ws, r, "perte de masse du PICA à 1250 K", 0.19, "[Wo] § 3.2 (balance ; 17 % à l'ATG)", key=True)
rpl = r - 1
r = kv(ws, r, "⇒ rendement en char de la résine DANS le PICA", f"=($C${rwr}-$C${rpl})/$C${rwr}",
       "(50 − 19)/50 = 0.62 — relevé par la capture de radicaux sur la préforme (H5)", font=F_CALC, nf="0.0000")
r = kv(ws, r, "couplage k = B'g/B'c du PICA", f"=$C${rpl}/(1-$C${rpl})",
       "k = m_gaz/m_char = 0.19/0.81 — calculé sur les MASSES", font=F_CALC, nf="0.0000")
rk = r - 1
r += 1

r = section(ws, r, "§8 — BLOC XML")
r = xmlblock(ws, r, [
    '<mixture thermo_db="NASA-9">',
    '    <species>',
    '       C H O N CH4 CN CO CO2 C2 C2H C2H2,acetylene C3 C4 C4H2,butadiyne C5',
    '       HCN H2 H2O N2 CH2OH CNN CNC CNCOCN C6H6 C6H5OH,phenol HNC C(gr)',
    '    </species>',
    '    <element_compositions default="air">',
    '        <composition name="air">N:0.79, O:0.21</composition>',
    '        <composition name="sc1008_pyro">C:0.2526, H:0.6407, O:0.1068</composition>',
    '        <composition name="sc1008_char">C:1.0</composition>',
    '    </element_compositions>',
    '</mixture>',
])
SC_XML = (rxml, rk)


# ===========================================================================
# ONGLET  LIÈGE / PHÉNOLIQUE
# ===========================================================================
ws = ws_cor
setup(ws, "BF8F00")
r = title(ws, "Liège / phénolique (cork phenolic) — 80 / 20 en masse",
          "Le SEUL matériau du dépôt dont le RENFORT PYROLYSE AUSSI : le gaz est le mélange de deux gaz, et le rapport renfort/résine entre dans la table B'",
          "data/mixtures/cork-air.xml · cork-pyrogas.xml")

r = section(ws, r, "§0 — FICHE D'IDENTITÉ")
r = kv(ws, r, "Matériau", "composite liège / résine phénolique, 80 % liège / 20 % résine en MASSE du solide", "matériau de type P50 (Amorim) — IXV et QARMAN")
r = kv(ws, r, "Renfort", "liège (Quercus suber) — subérine 45 / lignine 27 / polysaccharides 12 / tanins 6 / céroïdes 6 %", "IL PYROLYSE : c'est l'essentiel du gaz produit")
r = kv(ws, r, "Résine", "novolac C7H6O, rendement en char 50 %", "même motif que TACOT, PICA, CPh70")
r = kv(ws, r, "Rendement en char du composite", "20 % — mesuré par TGA sur le cork P50 (argon, 10 K/min, plateau 780-1650 K)", "Sakraker et al., CEAS Space Journal 14:377-393 (2022)")
r = kv(ws, r, "Sources", "cork_bprime/mise_en_donnees_cork.md · cork_pyrolysis_data.py", "")
r = kv(ws, r, "Compositions du XML", "air (-bl) · cork_pyro (-py) · cork_char (-char, -char-elem C)", "")
r += 1

r = section(ws, r, "§1 — HYPOTHÈSES")
r = hyp(ws, r, "H1", "LE RENFORT PYROLYSE. Le gaz est produit par les DEUX constituants, mélangés au prorata des masses réellement dégazées.",
        "C'est toute la différence avec un carbone/phénolique. Conséquence : le rapport liège/résine entre DANS la composition du gaz, donc dans la table B' — ce qui n'était pas le cas du CPh70.")
r = hyp(ws, r, "H2", "« 80 % de liège » se lit en MASSE, pas en volume.",
        "Piège d'énoncé : le liège est très léger (≈ 120 kg/m³ contre 1200 pour la résine). 80 % en VOLUME ne ferait que 28.6 % en masse — un tout autre matériau (gaz C:0.277, char 39.3 %, k = 1.545).")
r = hyp(ws, r, "H3", "Analyse élémentaire du liège retenue : C 62.4 / H 8.5 / O 28.4 % masse (littérature, sec sans cendres).",
        "PROVENANCE À CONFIRMER : valeurs usuelles pour Quercus suber, non vérifiées sur une source primaire accessible. Le contrôle biochimique du §5 est, lui, traçable ligne à ligne. L'azote (~0.6 %) est négligé — le retenir ajouterait une 4e composante au gaz.")
r = hyp(ws, r, "H4", "Rendement en char du liège DÉDUIT du composite mesuré et de la résine : 0.80·y + 0.20·0.50 = 0.20 ⇒ y = 12.5 %.",
        "La TGA ne sépare pas les deux constituants. Lever l'hypothèse « résine à 50 % » demanderait une seconde TGA, sur la résine seule ou sur le liège seul.")
r = hyp(ws, r, "H5", "Les deux constituants carbonisent vers du carbone pur ⇒ char C:1.0.",
        "Un char de liège réel retient un peu de H et de O : le §5 chiffre cette variante (char multi-élément C:0.864, H:0.108, O:0.027) — et rappelle qu'il faut alors RETIRER ces atomes du gaz.")
r = hyp(ws, r, "H6", "Le couplage stationnaire se calcule sur les MASSES : k = (1 − y)/y = 4.0.",
        "L'identité k = (ρ_v − ρ_c)/ρ_c suppose un volume constant, que ce matériau ne respecte PAS : le char se rétracte (V_char/V_vierge ≈ 0.32). Les densités donneraient k = 0.61, six fois trop peu.")
r += 1

r = section(ws, r, "§2 — DONNÉES D'ENTRÉE")
r = step(ws, r, "Analyse élémentaire du liège — valeurs de littérature (RETENUES)", "% masse, sec sans cendres")
rcl = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "", ""])
r = row(ws, r, ["liège [% masse]", 62.4, 8.5, 28.4, f"=SUM(C{r}:E{r})", "", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_CALC, None, None],
        nfs=[None, NF_P, NF_P, NF_P, NF_P], fills=[None, FILL_KEY, FILL_KEY, FILL_KEY, None])
r = note(ws, r, "La somme vaut 99.3 % et non 100 : l'étape 1 renormalise. C'est exactement ce que fait "
                "cork_pyrolysis_data.py.")
r += 1
r = step(ws, r, "Constituants biochimiques du liège et leurs unités de répétition (sert au contrôle du §5)",
         "M_i = Σ_E ν_E,i · M_E")
rbio = r + 1
r = headrow(ws, r, ["Constituant — unité de répétition", "% masse", "ν_C", "ν_H", "ν_O",
                    "M [g/mol]", "rendement en char"])
BIO = [("subérine — C57H104O12 (glycérol + 3 acides époxy-hydroxy C18, − 3 H2O)", 45.0, 57, 104, 12, 0.15),
       ("lignine — C10H12O3 (alcool coniférylique, unité guaïacyle)", 27.0, 10, 12, 3, 0.45),
       ("polysaccharides — C6H10O5 (anhydroglucose)", 12.0, 6, 10, 5, 0.15),
       ("tanins — C15H14O6 (flavan-3-ol, catéchine)", 6.0, 15, 14, 6, 0.40),
       ("céroïdes — C30H50O (friedéline)", 6.0, 30, 50, 1, 0.02)]
for lab, pct, nc, nh, no, cy in BIO:
    r = row(ws, r, [lab, pct, nc, nh, no,
                    f"=D{r}*{MC}+E{r}*{MH}+F{r}*{MO}", cy],
            fonts=[F_LAB, F_IN, F_IN, F_IN, F_IN, F_CALC, F_IN],
            nfs=[None, NF_P, "0", "0", "0", "0.00", "0.00"])
rbio_end = r - 1
r = row(ws, r, ["SOMME des parts", f"=SUM(C{rbio}:C{rbio_end})", "", "", "", "", ""],
        fonts=[F_LABB, F_CALC, None, None, None, None, None], nfs=[None, NF_P])
rbio_sum = r - 1
r = note(ws, r, "Les parts somment à 96 % : le complément (cendres ~1 %, humidité résiduelle, extractibles "
                "mineurs) n'est pas de la matière C/H/O identifiée. On renormalise donc sur les 96 % déclarés. "
                "Les rendements en char par constituant (dernière colonne) sont des ordres de grandeur d'ATG "
                "lente sous inerte — ils servent au contrôle B du §5, pas au calcul du XML.")
r += 1
r = kv(ws, r, "fraction massique de liège dans le solide", 0.80, "hypothèse H2 — lecture MASSIQUE", key=True)
rwc = r - 1
r = kv(ws, r, "fraction massique de résine dans le solide", f"=1-C{rwc}", "", font=F_CALC, nf=NF_X3)
rwr = r - 1
r = kv(ws, r, "rendement en char du COMPOSITE (mesuré, TGA P50)", 0.20, "plateau de 780 à 1650 K, argon, 10 K/min", key=True)
ryco = r - 1
r = kv(ws, r, "rendement en char de la RÉSINE", 0.50, "valeur classique novolac (énoncé) — même valeur que le TACOT", key=True)
ryr = r - 1
r = kv(ws, r, "motif de la résine — ν_C", 7, "novolac linéaire C7H6O, cf. Constantes §3")
rnu = r - 1
r = kv(ws, r, "                    ν_H", 6, "")
r = kv(ws, r, "                    ν_O", 1, "")
r = kv(ws, r, "ρ vierge / ρ char mesurés du P50 [kg/m³]", 465.6, "Sakraker et al. : 464.5 et 466.7 vierge")
rrv = r - 1
r = kv(ws, r, "", 289.1, "Sakraker et al. : 279.9 et 298.4 char — n'entre PAS dans k, cf. H6")
rrc = r - 1
r += 1

r = section(ws, r, "§3 — CALCUL PAS À PAS")
r = step(ws, r, "Étape 1 — renormaliser l'analyse élémentaire du liège à 100 %",
         "w_E = w_E,brut / Σ w_j,brut")
rw = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "", ""])
r = row(ws, r, ["liège renormalisé [% masse]", f"=C{rcl}/$F${rcl}*100", f"=D{rcl}/$F${rcl}*100",
                f"=E{rcl}/$F${rcl}*100", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_P, NF_P, NF_P, NF_P])
rwn = r - 1
r = step(ws, r, "Étape 2 — déduire le rendement en char du LIÈGE du rendement mesuré du composite",
         "w_liège·y_liège + w_résine·y_résine = y_composite")
r = kv(ws, r, "y_liège = (y_comp − w_rés·y_rés) / w_liège",
       f"=($C${ryco}-$C${rwr}*$C${ryr})/$C${rwc}",
       "attendu 0.125 — c'est LE nombre déduit, et le plus incertain de la chaîne (§6)",
       font=F_CALC, nf=NF_X3)
ryco_l = r - 1
r = step(ws, r, "Étape 3 — bilan de masse sur 100 g de composite vierge",
         "m_char,i = m_i · y_i ;  m_gaz,i = m_i · (1 − y_i)")
rbm = r + 1
r = headrow(ws, r, ["Constituant", "masse [g]", "rendement char", "char [g]", "gaz [g]", "", ""])
r = row(ws, r, ["liège", f"=$C${rwc}*100", f"=$C${ryco_l}", f"=C{r}*D{r}", f"=C{r}-E{r}", "", ""],
        nfs=[None, "0.00", NF_X3, "0.00", "0.00"])
rbc = r - 1
r = row(ws, r, ["résine", f"=$C${rwr}*100", f"=$C${ryr}", f"=C{r}*D{r}", f"=C{r}-E{r}", "", ""],
        nfs=[None, "0.00", NF_X3, "0.00", "0.00"])
rbr = r - 1
r = row(ws, r, ["COMPOSITE", f"=C{rbc}+C{rbr}", f"=(E{rbc}+E{rbr})/(C{rbc}+C{rbr})",
                f"=E{rbc}+E{rbr}", f"=F{rbc}+F{rbr}", "", ""],
        fonts=[F_LABB, F_CALC, F_CALC, F_CALC, F_CALC, None, None],
        nfs=[None, "0.00", NF_X3, "0.00", "0.00"])
rbt = r - 1
r = note(ws, r, "Contrôle : la colonne « rendement char » du composite doit redonner exactement les 20 % "
                "mesurés — c'est la définition même de l'étape 2.")
r = step(ws, r, "Étape 4 — fermeture élémentaire, constituant par constituant : aucun atome n'est créé",
         "n_E(gaz, i) = n_E(constituant, i) − n_E(char, i)")
rfe = r + 1
r = headrow(ws, r, ["Moles d'atomes", "C", "H", "O", "Somme", "x_E du gaz partiel", ""])
r = row(ws, r, ["liège — atomes du constituant",
                f"=C{rbc}*C{rwn}/100/{MC}", f"=C{rbc}*D{rwn}/100/{MH}", f"=C{rbc}*E{rwn}/100/{MO}",
                f"=SUM(C{r}:E{r})", "", ""], nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rlc = r - 1
r = row(ws, r, ["liège — atomes du char (C pur)", f"=E{rbc}/{MC}", 0, 0, f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rlch = r - 1
r = row(ws, r, ["liège — GAZ = constituant − char", f"=C{rlc}-C{rlch}", f"=D{rlc}-D{rlch}",
                f"=E{rlc}-E{rlch}", f"=SUM(C{r}:E{r})",
                "C:0.290 H:0.587 O:0.124", ""],
        fonts=[F_LAB, F_CALC, F_CALC, F_CALC, F_CALC, F_NOTE, None],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL, None])
rlg = r - 1
r = step(ws, r, "Étape 5 — même fermeture pour la résine (motif C7H6O)",
         "n_E = m_résine·ν_E / M_motif ;  M_motif = Σ ν_E·M_E")
r = kv(ws, r, "M du motif de résine [g/mol]",
       f"=C{rnu}*{MC}+C{rnu+1}*{MH}+C{rnu+2}*{MO}", "attendu 106.12 g/mol", font=F_CALC, nf=NF_M)
rMr = r - 1
rfr = r + 1
r = headrow(ws, r, ["Moles d'atomes", "C", "H", "O", "Somme", "x_E du gaz partiel", ""])
r = row(ws, r, ["résine — atomes du constituant",
                f"=C{rbr}/$C${rMr}*$C${rnu}", f"=C{rbr}/$C${rMr}*$C${rnu+1}", f"=C{rbr}/$C${rMr}*$C${rnu+2}",
                f"=SUM(C{r}:E{r})", "", ""], nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rrc2 = r - 1
r = row(ws, r, ["résine — atomes du char (C pur)", f"=E{rbr}/{MC}", 0, 0, f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rrch = r - 1
r = row(ws, r, ["résine — GAZ = constituant − char", f"=C{rrc2}-C{rrch}", f"=D{rrc2}-D{rrch}",
                f"=E{rrc2}-E{rrch}", f"=SUM(C{r}:E{r})", "C:0.269 H:0.626 O:0.104", ""],
        fonts=[F_LAB, F_CALC, F_CALC, F_CALC, F_CALC, F_NOTE, None],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL, None])
rrg = r - 1
r = step(ws, r, "Étape 6 — mélanger les deux gaz et normaliser",
         "n_E(gaz) = n_E(gaz, liège) + n_E(gaz, résine) ;  x_E = n_E / Σ n_j")
rmix = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "M_gaz [g/mol]", "H/O"])
r = row(ws, r, ["gaz du composite — moles d'atomes", f"=C{rlg}+C{rrg}", f"=D{rlg}+D{rrg}",
                f"=E{rlg}+E{rrg}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rmn = r - 1
r = row(ws, r, ["x_E (fraction molaire élémentaire)", f"=C{rmn}/$F${rmn}", f"=D{rmn}/$F${rmn}",
                f"=E{rmn}/$F${rmn}", f"=SUM(C{r}:E{r})",
                f"=C{r}*{MC}+D{r}*{MH}+E{r}*{MO}", f"=D{r}/E{r}"],
        nfs=[None, NF_X, NF_X, NF_X, NF_X3, NF_M, "0.000"])
rxE = r - 1
r = row(ws, r, ["y_E (fraction massique élémentaire)", f"=C{rxE}*{MC}/$G${rxE}", f"=D{rxE}*{MH}/$G${rxE}",
                f"=E{rxE}*{MO}/$G${rxE}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_X3, NF_X3, NF_X3, NF_X3])
r = note(ws, r, "H/O = 4.89 : bien plus bas que les 5.9-6.8 des phénoliques. C'est normal et attendu — "
                "le liège (subérine, lignine, polysaccharides) est nativement bien plus oxygéné qu'un "
                "phénolique. L'invariant H/O de l'onglet SC-1008 ne s'applique donc PAS à ce matériau.")
r += 1

r = section(ws, r, "§4 — RÉSULTAT : CE QUI ENTRE DANS LE XML")
r, rcalc, rxml = result_block(ws, r, "cork_pyro", f"C{rxE}", f"D{rxE}", f"E{rxE}",
                              (0.287, 0.592, 0.121))
r = headrow(ws, r, ["Char (cork_char)", "C", "H", "O", "", "", ""])
r = row(ws, r, ["composition du XML", 1.0, "—", "—", "", "", ""],
        fonts=[F_LAB, F_OUT, F_NOTE, F_NOTE, None, None, None],
        nfs=[None, NF_X3, None, None], fills=[None, FILL_OUT, None, None])
r += 1

r = section(ws, r, "§5 — CONTRÔLES ET VALIDATION CROISÉE")
r = step(ws, r, "Contrôle A — reconstruire le liège depuis ses unités de répétition (traçable ligne à ligne)",
         "w_E = Σ_i (part_i / Σ parts) · (ν_E,i·M_E / M_i)")
rrec = r + 1
r = headrow(ws, r, ["", "C", "H", "O", "Somme", "écart / littérature [points]", ""])
r = row(ws, r, ["liège reconstruit [% masse]",
                f"=SUMPRODUCT($C${rbio}:$C${rbio_end},D{rbio}:D{rbio_end}/$G${rbio}:$G${rbio_end})*{MC}/$C${rbio_sum}*100",
                f"=SUMPRODUCT($C${rbio}:$C${rbio_end},E{rbio}:E{rbio_end}/$G${rbio}:$G${rbio_end})*{MH}/$C${rbio_sum}*100",
                f"=SUMPRODUCT($C${rbio}:$C${rbio_end},F{rbio}:F{rbio_end}/$G${rbio}:$G${rbio_end})*{MO}/$C${rbio_sum}*100",
                f"=SUM(C{r}:E{r})",
                f"=MAX(ABS(C{r}-C{rwn}),ABS(D{r}-D{rwn}),ABS(E{r}-E{rwn}))", ""],
        nfs=[None, NF_P, NF_P, NF_P, NF_P, NF_P])
r = note(ws, r, "Attendu C 66.16 / H 8.71 / O 25.13. L'hydrogène tombe à 0.2 point près ; le carbone est "
                "surestimé d'environ 4 points et l'oxygène sous-estimé d'autant. Deux causes non exclusives : "
                "(a) les 4 % manquants, s'ils sont de la matière oxygénée, rapprocheraient la reconstruction "
                "à ~1 point ; (b) l'unité choisie pour la subérine, qui pèse 45 % à elle seule. "
                "Effet sur le gaz : C:0.300 / H:0.594 / O:0.107 au lieu de C:0.287 / H:0.592 / O:0.121 — "
                "réel, mais du second ordre devant l'incertitude sur le rendement en char (§6).")
r = step(ws, r, "Contrôle B — rendement en char du liège par additivité des constituants",
         "y_liège = Σ_i part_i · y_i / Σ parts")
r = kv(ws, r, "y_liège par additivité",
       f"=SUMPRODUCT($C${rbio}:$C${rbio_end},$H${rbio}:$H${rbio_end})/$C${rbio_sum}",
       "attendu 24.2 % — contre 12.5 % déduit de la TGA du composite", font=F_CALC, nf=NF_X3)
rya = r - 1
r = kv(ws, r, "⇒ rendement du composite impliqué", f"=$C${rwc}*C{rya}+$C${rwr}*$C${ryr}",
       "29.3 % contre 20.0 % mesurés sur le P50 : les deux voies NE SE REJOIGNENT PAS, d'un facteur deux sur le liège",
       font=F_CALC, nf=NF_X3)
r = note(ws, r, "On garde la MESURE. Explications de l'écart : le P50 est plastifié au glycol, qui part "
                "entièrement en gaz ; son rapport liège/résine n'est pas nécessairement 80/20 ; et les "
                "rendements par constituant ci-dessus sont indicatifs. C'est la mesure qui fixe k, "
                "la grandeur qui gouverne le point de fonctionnement.")
r = step(ws, r, "Contrôle C — variante : char de liège retenant H et O (char multi-élément)",
         "n_E(char) = Σ_i (m_char,i / M_i) · ν_E,i")
rmul = r + 1
r = headrow(ws, r, ["char de liège supposé [% masse]", "C", "H", "O", "Somme", "", ""])
r = row(ws, r, ["hypothèse de travail", 90.0, 2.0, 8.0, f"=SUM(C{r}:E{r})", "", ""],
        fonts=[F_LAB, F_IN, F_IN, F_IN, F_CALC, None, None],
        nfs=[None, NF_P, NF_P, NF_P, NF_P])
rmulh = r - 1
r = row(ws, r, ["moles d'atomes du char total (liège + résine)",
                f"=E{rbc}*C{rmulh}/100/{MC}+E{rbr}/{MC}", f"=E{rbc}*D{rmulh}/100/{MH}",
                f"=E{rbc}*E{rmulh}/100/{MO}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_MOL, NF_MOL, NF_MOL, NF_MOL])
rmulm = r - 1
r = row(ws, r, ["⇒ cork_char multi-élément (x_E)", f"=C{rmulm}/$F${rmulm}", f"=D{rmulm}/$F${rmulm}",
                f"=E{rmulm}/$F${rmulm}", f"=SUM(C{r}:E{r})", "", ""],
        nfs=[None, NF_X3, NF_X3, NF_X3, NF_X3])
r = note(ws, r, "Attendu C:0.864, H:0.108, O:0.027. `-char-elem C` reste valable : l'élément du bilan doit "
                "être présent dans le char et absent (ou minoritaire) au bord de couche limite. ATTENTION À LA "
                "COHÉRENCE : si le char retient des atomes, il faut les RETIRER du gaz (le gaz devient "
                "C:0.297, H:0.585, O:0.119). C'est automatique dans la fermeture élémentaire, pas si l'on "
                "recopie deux compositions de sources différentes.")
r += 1

r = section(ws, r, "§6 — SENSIBILITÉ")
r = step(ws, r, "Au rendement en char du liège — LE paramètre à mesurer en priorité", "")
rs1 = r + 1
r = headrow(ws, r, ["y_liège testé", "C", "H", "O", "y composite impliqué", "k = (1−y)/y", ""])
for y in [0.05, 0.10, 0.125, 0.15, 0.20, 0.242]:
    cC = f"($C${rlc}-$C${rbc}*B{r}/{MC}+$C${rrg})"
    cH = f"($D${rlc}+$D${rrg})"
    cO = f"($E${rlc}+$E${rrg})"
    tot = f"({cC}+{cH}+{cO})"
    r = row(ws, r, [y, f"={cC}/{tot}", f"={cH}/{tot}", f"={cO}/{tot}",
                    f"=$C${rwc}*B{r}+$C${rwr}*$C${ryr}", f"=(1-F{r})/F{r}", ""],
            fonts=[F_IN, F_CALC, F_CALC, F_CALC, F_CALC, F_CALC, None],
            nfs=[NF_X3, NF_X3, NF_X3, NF_X3, NF_X3, "0.000"])
r = note(ws, r, "±5 points de rendement déplacent le carbone du gaz de ±0.018 et k de ∓25 %. "
                "Priorité de mesure : la TGA d'abord, l'analyse élémentaire du liège ensuite — "
                "k gouverne le point de fonctionnement, la composition ne le déplace que marginalement.")
r = step(ws, r, "Au rapport liège / résine — ce que le mélange change VRAIMENT", "")
rs2 = r + 1
r = headrow(ws, r, ["% masse de liège", "C", "H", "O", "y composite", "k = B'g/B'c", ""])
for w in [0.0, 0.5, 0.8, 1.0]:
    mc_ = f"(B{r}*100)"
    mr_ = f"((1-B{r})*100)"
    gC = f"({mc_}*$C${rwn}/100/{MC}-{mc_}*$C${ryco_l}/{MC}+{mr_}/$C${rMr}*$C${rnu}-{mr_}*$C${ryr}/{MC})"
    gH = f"({mc_}*$D${rwn}/100/{MH}+{mr_}/$C${rMr}*$C${rnu+1})"
    gO = f"({mc_}*$E${rwn}/100/{MO}+{mr_}/$C${rMr}*$C${rnu+2})"
    tt = f"({gC}+{gH}+{gO})"
    r = row(ws, r, [w, f"={gC}/{tt}", f"={gH}/{tt}", f"={gO}/{tt}",
                    f"=B{r}*$C${ryco_l}+(1-B{r})*$C${ryr}", f"=(1-F{r})/F{r}", ""],
            fonts=[F_IN, F_CALC, F_CALC, F_CALC, F_CALC, F_CALC, None],
            nfs=[NF_X3, NF_X3, NF_X3, NF_X3, NF_X3, "0.000"])
r = note(ws, r, "Conclusion honnête : le rapport liège/résine entre bien dans la table B' (contrairement au "
                "cas carbone/phénolique), mais son effet DIRECT sur la composition du gaz est modeste — les "
                "deux gaz se ressemblent. Son effet DOMINANT est sur la QUANTITÉ de gaz, c.-à-d. sur k, "
                "multiplié par sept entre la résine pure et le liège pur.")
r += 1

r = section(ws, r, "§7 — RÉPONSE MATÉRIAU (n'entre PAS dans le XML)")
r = kv(ws, r, "k = B'g/B'c = m_gaz / m_char = (1 − y)/y", f"=F{rbt}/E{rbt}",
       "attendu 4.0 — QUINZE FOIS le TACOT (0.273). Le matériau fonctionne à très fort soufflage pyrolytique.",
       font=F_CALC, nf="0.0000")
rk = r - 1
r = kv(ws, r, "k si l'on prenait les densités : (ρ_v − ρ_c)/ρ_c", f"=(C{rrv}-C{rrc})/C{rrc}",
       "0.61 — SIX FOIS TROP PEU. Le piège de H6 : l'identité suppose un volume constant, que le P50 ne respecte pas.",
       font=F_CALC, nf="0.0000")
r = kv(ws, r, "V_char / V_vierge impliqué", f"=$C${ryco}*C{rrv}/C{rrc}",
       "≈ 0.32 alors que la masse résiduelle est 0.20 : le char se rétracte", font=F_CALC, nf=NF_X3)
r = note(ws, r, "Conséquences pratiques : (1) balayer B'g jusqu'à ~10, pas jusqu'à 2 — une table s'arrêtant "
                "à B'g = 2 ne couvre pas le point de fonctionnement au-delà de ~3500 K ; (2) la récession se "
                "lit sur la ρ_char MESURÉE (289.1 kg/m³) et non sur 0.20·ρ_v = 93 kg/m³, sans quoi elle serait "
                "surestimée d'un facteur trois ; (3) le soufflage divise par 2.6 l'ablation du char en régime "
                "d'oxydation à 1000 K — effet marginal pour un TACOT, dominant ici.")
r += 1

r = section(ws, r, "§8 — BLOC XML")
r = xmlblock(ws, r, [
    '<mixture thermo_db="NASA-9">',
    '    <species>',
    '       C H O N CH4 CN CO CO2 C2 C2H C2H2,acetylene C3 C4 C4H2,butadiyne C5',
    '       HCN H2 H2O N2 CH2OH CNN CNC CNCOCN C6H6 HNC C(gr)',
    '    </species>',
    '    <element_compositions default="air">',
    '        <composition name="air">N:0.79, O:0.21</composition>',
    '        <composition name="cork_pyro">C:0.287, H:0.592, O:0.121</composition>',
    '        <composition name="cork_char">C:1.0</composition>',
    '    </element_compositions>',
    '</mixture>',
])
COR_XML = (rxml, rk)


# ===========================================================================
# ONGLET  CARBONE (graphite)
# ===========================================================================
ws = ws_car
setup(ws, "404040")
r = title(ws, "Carbone graphite — ablateur non pyrolysant",
          "Cas dégénéré : le solide EST le char, il n'y a pas de gaz de pyrolyse · sert de cas de référence physique aux autres onglets",
          "data/mixtures/carbon-air.xml")

r = section(ws, r, "§0 — FICHE D'IDENTITÉ")
r = kv(ws, r, "Matériau", "graphite pur (carbone massif)", "carbon_bprime/bprime_carbon_physique.md")
r = kv(ws, r, "Renfort / résine", "sans objet — matériau monolithique", "aucune pyrolyse interne")
r = kv(ws, r, "Système chimique", "C-N-O : 5 espèces d'air + 3 formes gazeuses du carbone + CO, CO2, CN + C(gr)", "12 espèces au total")
r = kv(ws, r, "Compositions du XML", "air (-bl) · pyro = C:1.0 (-py, et définit le char)", "")
r += 1

r = section(ws, r, "§1 — HYPOTHÈSES")
r = hyp(ws, r, "H1", "Pas de pyrolyse interne : B'g = 0.",
        "Il n'y a pas de résine à décomposer. Toute la masse injectée vient de la surface, d'où le nom de « cas dégénéré ».")
r = hyp(ws, r, "H2", "Le char est le matériau lui-même : carbone pur, C:1.0.",
        "Aucune fermeture élémentaire à faire — c'est la seule mise en données du dépôt où la composition ne se CALCULE pas.")
r = hyp(ws, r, "H3", "Équilibre thermochimique à la paroi, char infini.",
        "Le solveur ajoute δ = max(200, 100·B'g) de carbone élémentaire à la composition du bord de couche limite avant de normaliser, puis calcule l'équilibre multiphase.")
r = hyp(ws, r, "H4", "C(gr) est indispensable dans la liste d'espèces.",
        "Sans phase condensée, C3 gazeux domine dès 300 K, Y_w,C → 1 et B'c ≈ 200 — non physique. C'est le contrôle du §5, valable pour TOUS les onglets carbonés.")
r += 1

r = section(ws, r, "§2 — DONNÉES D'ENTRÉE")
r = kv(ws, r, "bord de couche limite", "air : N:0.79, O:0.21 (molaire)", "cf. onglet Constantes §2")
r = kv(ws, r, "fraction massique d'oxygène de l'air", "=Constantes!D19", "y_O calculé au §2 de l'onglet Constantes",
       font=F_LINK, nf=NF_X)
ryO = r - 1
r = kv(ws, r, "masse volumique du graphite [kg/m³]", 1800, "ordre de grandeur (1600-2200 selon la qualité) — n'entre que dans la récession")
rrho = r - 1
r += 1

r = section(ws, r, "§3 — CALCUL PAS À PAS")
r = step(ws, r, "Étape 1 — composition élémentaire du char : directe",
         "le solide est du carbone pur ⇒ x_C = 1")
rc = r + 1
r = headrow(ws, r, ["", "C", "N", "O", "Somme", "", ""])
r = row(ws, r, ["char (= le matériau)", 1.0, 0, 0, f"=SUM(C{r}:E{r})", "", ""],
        fonts=[F_LAB, F_OUT, F_CALC, F_CALC, F_CALC, None, None],
        nfs=[None, NF_X3, NF_X3, NF_X3, NF_X3], fills=[None, FILL_OUT, None, None, None])
r = step(ws, r, "Étape 2 — gaz de pyrolyse : il n'y en a pas",
         "B'g = 0 dans toutes les commandes bprime de ce matériau")
r = para(ws, r, "La composition nommée `pyro` du XML vaut néanmoins C:1.0. Ce n'est pas une incohérence : "
                "elle sert d'argument à `-py` (le parseur exige une composition valide) et définit le char. "
                "Comme elle est appelée avec -b 0.0, elle ne contribue à rien.", 30)
r = step(ws, r, "Étape 3 — le bilan de surface se réduit à une expression fermée",
         "B'c = [Y_e,C + B'g·Y_g,C − Y_w,C(1+B'g)] / (Y_w,C − Y_c,C)")
r = para(ws, r, "Avec Y_c,C = 1 (char de carbone pur), Y_e,C = 0 (l'air ne contient pas de carbone) "
                "et B'g = 0, il ne reste que :", 16)
r = xmlblock(ws, r, ["   B'c = − Y_w,C / (Y_w,C − 1) = Y_w,C / (1 − Y_w,C)"])
r = para(ws, r, "La seule inconnue est Y_w,C, la fraction massique de carbone dans le gaz d'équilibre "
                "à la paroi — c'est elle que le solveur multiphase calcule à (T_w, P).", 30)
r += 1

r = section(ws, r, "§4 — RÉSULTAT : CE QUI ENTRE DANS LE XML")
rres = r + 1
r = headrow(ws, r, ["Composition élémentaire (fractions molaires)", "C", "N", "O", "Somme", "", ""])
r = row(ws, r, ["pyro / char — valeur portée dans le XML", 1.0, "—", "—", "", "", ""],
        fonts=[F_LAB, F_OUT, F_NOTE, F_NOTE, None, None, None],
        nfs=[None, NF_X3, None, None], fills=[None, FILL_OUT, None, None])
r = row(ws, r, ["air — valeur portée dans le XML", "—", 0.79, 0.21, f"=D{r}+E{r}", "", ""],
        fonts=[F_LAB, F_NOTE, F_OUT, F_OUT, F_CALC, None, None],
        nfs=[None, None, NF_X3, NF_X3, NF_X3], fills=[None, None, FILL_OUT, FILL_OUT, None])
r += 1

r = section(ws, r, "§5 — CONTRÔLES ET VALIDATION CROISÉE")
r = step(ws, r, "Contrôle de recette valable pour TOUS les matériaux carbonés du dépôt",
         "à 300 K, B'g = 0 : régime d'oxydation basse température, C(gr) + O2 → CO2")
r = para(ws, r, "À 300 K tout l'oxygène de l'air est consommé en CO2. Chaque O2 (2 atomes d'oxygène) "
                "emporte un atome de carbone : la masse de carbone ablatée par unité de masse d'air vaut "
                "donc y_O · M_C / (2·M_O). Ce nombre ne dépend PAS de la pression.", 30)
r = kv(ws, r, "B'c attendu à 300 K, 1 atm", f"=C{ryO}*{MC}/(2*{MO})",
       "attendu 0.0874 — valeur obtenue par bprime sur carbon-air, cork-air, tacot-air…", font=F_CALC, nf="0.00000")
r = kv(ws, r, "B'c obtenu si C(gr) manque dans <species>", "≈ 200",
       "signature sans ambiguïté d'une liste d'espèces incomplète : Y_w,C → 1 et B'c sature sur son plafond interne",
       font=F_CALC)
r += 1

r = section(ws, r, "§6 — SENSIBILITÉ — régimes physiques (à 1 atm)")
rrg = r + 1
r = headrow(ws, r, ["Régime", "T_w [K]", "Phase stable", "Espèce gazeuse dominante", "Mécanisme", "", ""])
for lab, T, ph, sp, me in [
        ("faible ablation", "300 – 700", "C(gr)", "CO2, O2, N2", "C(gr) + O2 → CO2"),
        ("oxydation active", "700 – 3500", "C(gr)", "CO, N2, CN", "C(gr) + O → CO ; Boudouard C(gr) + CO2 → 2 CO"),
        ("sublimation", "3500 – 4000", "C(gr) + gaz", "CO, CN, C3", "C(gr) ⇌ C, C2, C3"),
        ("sublimation totale", "> 4000", "gaz seul", "C, C2, C3", "plus de phase condensée : B'c ≫ 1")]:
    r = row(ws, r, [lab, T, ph, sp, me, "", ""],
            fonts=[F_LAB, F_IN, F_IN, F_IN, F_NOTE, None, None])
r = note(ws, r, "À haute pression les mêmes régimes sont décalés vers des températures plus élevées "
                "(la sublimation est retardée, effet Clausius-Clapeyron), tandis que les réactions "
                "d'oxydation sont favorisées. La table B' est donc une SURFACE B'c(T_w, P), tracée comme "
                "un faisceau de 25 isobares de 10⁻³ à 10³ atm.")
r += 1

r = section(ws, r, "§7 — RÉPONSE MATÉRIAU (n'entre PAS dans le XML)")
r = kv(ws, r, "couplage k = B'g / B'c", 0.0, "aucun gaz de pyrolyse — le graphite est le seul matériau du classeur à k = 0",
       font=F_CALC, nf=NF_X3)
rk = r - 1
r = kv(ws, r, "récession par unité de flux, B'c/ρ_c à 300 K [m³/kg]",
       f"=C{ryO}*{MC}/(2*{MO})/C{rrho}", "ṡ = B'c·ṁe/ρ_c", font=F_CALC, nf="0.00E+00")
r += 1

r = section(ws, r, "§8 — BLOC XML")
r = xmlblock(ws, r, [
    '<mixture thermo_db="NASA-9">',
    '    <species>',
    '        N O NO N2 O2 C C2 C3 CN CO CO2 C(gr)',
    '    </species>',
    '    <element_compositions default="air">',
    '        <composition name="air">N:0.79, O:0.21</composition>',
    '        <composition name="pyro">C:1.0</composition>',
    '    </element_compositions>',
    '</mixture>',
])
r += 1
r = para(ws, r, "bprime -T 300:25:5000 -P <P> -b 0.0 -m carbon-air -bl air -py pyro", 16)
CAR_XML = (rres, rk)


# ===========================================================================
# ONGLET  SILICE
# ===========================================================================
ws = ws_sil
setup(ws, "0070C0")
r = title(ws, "Silice SiO2 — char MULTI-ÉLÉMENT",
          "Le seul matériau du dépôt dont le char n'est pas du carbone : montre comment se remplit une composition à plusieurs éléments",
          "data/mixtures/silica-air.xml")

r = section(ws, r, "§0 — FICHE D'IDENTITÉ")
r = kv(ws, r, "Matériau", "silice pure SiO2", "silice_bprime/silice_bprime.py")
r = kv(ws, r, "Système chimique", "Si-N-O : 5 espèces d'air + 5 espèces gazeuses du silicium + 4 phases condensées de SiO2", "SiO2(a-qz), SiO2(b-qz), SiO2(b-crt), SiO2(L)")
r = kv(ws, r, "Binaire dédié", "bprime_silica (et non bprime)", "le bilan est résolu sur Si, pas sur C")
r = kv(ws, r, "Compositions du XML", "air (-bl) · silice (char, -char-elem Si)", "")
r += 1

r = section(ws, r, "§1 — HYPOTHÈSES")
r = hyp(ws, r, "H1", "Le char est la silice elle-même, de stœchiométrie SiO2 exacte.",
        "Matériau monolithique : pas de résine, donc pas de gaz de pyrolyse (B'g = 0) et pas de fermeture élémentaire à faire.")
r = hyp(ws, r, "H2", "L'élément du bilan est le SILICIUM, pas le carbone.",
        "Règle générale : `-char-elem` doit désigner un élément PRÉSENT dans le char et ABSENT (ou minoritaire) au bord de couche limite. L'oxygène ne conviendrait pas — l'air en contient 21 % molaire.")
r = hyp(ws, r, "H3", "Les quatre phases condensées de SiO2 doivent figurer dans <species>.",
        "Même exigence que C(gr) pour les matériaux carbonés : sans phase condensée, le solveur ne peut pas faire précipiter la silice et B'c sature sur une valeur non physique. Ici il faut les quatre polymorphes (quartz α et β, cristobalite β, liquide) pour couvrir toute la plage de température.")
r = hyp(ws, r, "H4", "Équilibre thermochimique à la paroi.", "Identique aux autres matériaux.")
r += 1

r = section(ws, r, "§2 — DONNÉES D'ENTRÉE")
r = kv(ws, r, "stœchiométrie — ν_Si", 1, "SiO2", key=True)
rnsi = r - 1
r = kv(ws, r, "stœchiométrie — ν_O", 2, "", key=True)
rno = r - 1
r = kv(ws, r, "masse volumique de la silice [kg/m³]", 2200, "silice vitreuse — n'entre que dans la récession")
rrho = r - 1
r += 1

r = section(ws, r, "§3 — CALCUL PAS À PAS")
r = step(ws, r, "Étape 1 — masse molaire du motif", "M = ν_Si·M_Si + ν_O·M_O")
r = kv(ws, r, "M(SiO2) [g/mol]", f"=C{rnsi}*{MSI}+C{rno}*{MO}", "attendu 60.08 g/mol", font=F_CALC, nf=NF_M)
rM = r - 1
r = step(ws, r, "Étape 2 — composition élémentaire, en molaire puis en massique",
         "x_E = ν_E / Σ ν_j ;  y_E = ν_E·M_E / M")
rcp = r + 1
r = headrow(ws, r, ["", "Si", "O", "Somme", "", "", ""])
r = row(ws, r, ["ν_E (moles d'atomes par motif)", f"=C{rnsi}", f"=C{rno}", f"=SUM(C{r}:D{r})", "", "", ""],
        nfs=[None, NF_X3, NF_X3, NF_X3])
rnu = r - 1
r = row(ws, r, ["x_E (fraction molaire élémentaire, normalisée)",
                f"=C{rnu}/$E${rnu}", f"=D{rnu}/$E${rnu}", f"=SUM(C{r}:D{r})", "", "", ""],
        nfs=[None, NF_X, NF_X, NF_X3])
rxE = r - 1
r = row(ws, r, ["y_E (fraction massique élémentaire)",
                f"=C{rnu}*{MSI}/$C${rM}", f"=D{rnu}*{MO}/$C${rM}", f"=SUM(C{r}:D{r})", "", "", ""],
        nfs=[None, NF_X, NF_X, NF_X3])
r = note(ws, r, "Attendu : x_Si = 0.3333 / x_O = 0.6667 en molaire, y_Si = 0.4674 / y_O = 0.5326 en masse. "
                "Le piège classique est de confondre les deux — le parseur suppose du MOLAIRE par défaut.")
r += 1

r = section(ws, r, "§4 — RÉSULTAT : CE QUI ENTRE DANS LE XML")
rres = r + 1
r = headrow(ws, r, ["Composition élémentaire", "Si", "O", "Somme", "après normalisation par le parseur", "", ""])
r = row(ws, r, ["valeur portée dans le XML (silice)", 1.0, 2.0, f"=SUM(C{r}:D{r})",
                "x_Si = 1/3, x_O = 2/3", ""],
        fonts=[F_LAB, F_OUT, F_OUT, F_CALC, F_NOTE, None],
        nfs=[None, NF_X3, NF_X3, NF_X3, None], fills=[None, FILL_OUT, FILL_OUT, None, None])
r = note(ws, r, "C'est l'illustration la plus nette de la règle de normalisation automatique "
                "(Composition::componentsFromList) : « Si:1.0, O:2.0 » somme à 3, et le parseur le ramène "
                "à 1. On peut donc saisir indifféremment des moles brutes, des pourcentages ou des "
                "fractions — c'est le RAPPORT qui compte, et écrire directement Si:0.3333, O:0.6667 "
                "donnerait exactement le même mélange.")
r += 1

r = section(ws, r, "§5 — CONTRÔLES ET VALIDATION CROISÉE")
r = kv(ws, r, "x_Si + x_O après normalisation", f"=C{rxE}+D{rxE}", "doit valoir 1 exactement", font=F_CALC, nf="0.000000")
r = kv(ws, r, "rapport O/Si", f"=D{rnu}/C{rnu}", "doit valoir 2 — la stœchiométrie de la silice", font=F_CALC, nf="0.000")
r = note(ws, r, "Le cas de la silice est aussi le CONTRE-EXEMPLE de référence pour les composites : dès que "
                "les constituants ne carbonisent pas vers le même produit (fibres de silice + résine "
                "phénolique, par exemple), le char devient multi-élément et il faut pondérer par les MASSES "
                "de char de chaque constituant : n_E(char) = Σ_i (m_char,i / M_i)·ν_E,i. Le ratio "
                "renfort/résine entre alors dans la table B' — cf. onglet CPh70 §6 pour le cas contraire, "
                "et l'onglet Liège-phénolique §5 pour un exemple chiffré de pondération.")
r += 1

r = section(ws, r, "§6 — SENSIBILITÉ")
r = para(ws, r, "Sans objet : la stœchiométrie SiO2 est exacte, il n'y a ni rendement en char, ni analyse "
                "élémentaire, ni motif de résine à choisir. C'est la mise en données la moins incertaine "
                "du classeur — l'incertitude est reportée sur la base thermodynamique des quatre phases "
                "condensées et sur l'hypothèse d'équilibre à la paroi.", 30)
r += 1

r = section(ws, r, "§7 — RÉPONSE MATÉRIAU (n'entre PAS dans le XML)")
r = kv(ws, r, "couplage k = B'g / B'c", 0.0, "pas de pyrolyse interne — comme le graphite", font=F_CALC, nf=NF_X3)
rk = r - 1
r = kv(ws, r, "ρ_char [kg/m³]", f"=C{rrho}", "la silice fondue est évacuée mécaniquement : la récession réelle dépasse celle donnée par B'c seul",
       font=F_CALC, nf="0")
r += 1

r = section(ws, r, "§8 — BLOC XML")
r = xmlblock(ws, r, [
    '<mixture thermo_db="NASA-9">',
    '    <species>',
    '        N O NO N2 O2 Si SiO SiO2 Si2 SiN SiO2(a-qz) SiO2(b-qz) SiO2(b-crt) SiO2(L)',
    '    </species>',
    '    <element_compositions default="air">',
    '        <composition name="air">N:0.79, O:0.21</composition>',
    '        <composition name="silice">Si:1.0, O:2.0</composition>',
    '    </element_compositions>',
    '</mixture>',
])
r += 1
r = para(ws, r, "bprime_silica -T 300:25:5000 -P <P> -b 0.0 -m silica-air -bl air", 16)
SIL_XML = (rres, rk)


# ===========================================================================
# ONGLET  SYNTHÈSE  (rempli en dernier, affiché en premier)
# ===========================================================================
ws = ws_syn
setup(ws, "1F3864")
r = title(ws, "Mise en données des matériaux — synthèse",
          "D'où viennent les proportions portées dans data/mixtures/*.xml : hypothèses, étapes et formules, un onglet par matériau",
          "data/mixtures/ — 8 matériaux, 16 fichiers de mélange")

r = section(ws, r, "§1 — MODE D'EMPLOI")
r = para(ws, r, "Chaque onglet matériau reconstruit, ÉTAPE PAR ÉTAPE ET EN FORMULES VIVANTES, la chaîne qui "
                "mène de la donnée expérimentale publiée aux trois nombres portés dans le fichier XML. "
                "Modifier une cellule bleue (donnée d'entrée) recalcule toute la chaîne jusqu'à la "
                "composition, et jusqu'à l'écart affiché avec la valeur réellement présente dans le dépôt.", 30)
r = para(ws, r, "Trame commune à tous les onglets : §0 fiche d'identité · §1 hypothèses · §2 données d'entrée "
                "(avec sources) · §3 calcul pas à pas · §4 résultat porté dans le XML · §5 contrôles et "
                "validation croisée · §6 sensibilité · §7 réponse matériau (hors XML) · §8 bloc XML.", 30)
r = para(ws, r, "L'onglet Constantes porte les masses molaires, l'air de couche limite, les motifs de résine "
                "et le rappel du bilan de masse de surface — il est référencé par tous les autres.", 16)
r += 1

r = section(ws, r, "§2 — TABLEAU DE SYNTHÈSE : LES COMPOSITIONS DES XML")
rsy = r + 1
r = headrow(ws, r, ["Matériau (onglet)", "gaz — C", "gaz — H", "gaz — O", "H/O", "char", "k = B'g/B'c"])
SYN = [
    ("TACOT",            "TACOT",            TAC_XML[0], TAC_XML[1], "C:1.0", True),
    ("CPh70",            "CPh70",            CPH_XML[0], CPH_XML[1], "C:1.0", True),
    ("ZURAM 18/50",      "ZURAM",            ZUR_XML[0], ZUR_XML[1], "C:1.0", True),
    ("SC-1008 (PICA)",   "SC-1008",          SC_XML[0],  SC_XML[1],  "C:1.0", True),
    ("Liège/phénolique", "Liège-phénolique", COR_XML[0], COR_XML[1], "C:1.0", True),
    ("Carbone graphite", "Carbone",          CAR_XML[0], CAR_XML[1], "C:1.0", False),
    ("Silice SiO2",      "Silice",           SIL_XML[0], SIL_XML[1], "Si:1.0, O:2.0", False),
]
for lab, sheet, rx, rk_, char, has_gas in SYN:
    q = f"'{sheet}'"
    if has_gas:
        vals = [lab, f"={q}!C{rx}", f"={q}!D{rx}", f"={q}!E{rx}",
                f"={q}!D{rx}/{q}!E{rx}", char, f"={q}!C{rk_}"]
        fts = [F_LAB, F_LINK, F_LINK, F_LINK, F_CALC, F_OUT, F_LINK]
        nfs = [None, NF_X, NF_X, NF_X, "0.00", None, "0.0000"]
    else:
        vals = [lab, "—", "—", "—", "—", char, f"={q}!C{rk_}"]
        fts = [F_LAB, F_NOTE, F_NOTE, F_NOTE, F_NOTE, F_OUT, F_LINK]
        nfs = [None, None, None, None, None, None, "0.0000"]
    r = row(ws, r, vals, fonts=fts, nfs=nfs)
r = note(ws, r, "Les colonnes « gaz » sont des liens vers la ligne « valeur portée dans le XML » de chaque "
                "onglet. Le graphite et la silice n'ont pas de gaz de pyrolyse (k = 0). "
                "TACOT et CPh70 portent RIGOUREUSEMENT la même composition : leurs fichiers XML sont "
                "identiques, seul k les distingue — c'est la démonstration de l'onglet CPh70.")
r += 1

r = section(ws, r, "§3 — LES TROIS SEULES DONNÉES QUE CONSOMME LE CALCUL")
rtr = r + 1
r = headrow(ws, r, ["#", "Composition", "Rôle dans le bilan", "Option", "Comment on l'obtient", "", ""])
for n, comp, role, opt, how in [
        (1, "bord de couche limite", "Y_e", "-bl", "l'atmosphère : air N:0.79 O:0.21 · Mars · Titan…"),
        (2, "gaz de pyrolyse", "Y_g", "-py", "c'est TOUT l'objet de ce classeur — quatre routes possibles, §4"),
        (3, "char", "Y_c", "-char + -char-elem", "C:1.0 si tous les constituants carbonisent vers le carbone ; sinon pondérer par les masses de char")]:
    r = row(ws, r, [n, comp, role, opt, how, "", ""],
            fonts=[F_LABB, F_LAB, F_LAB, F_MONO, F_NOTE, None, None])
r = note(ws, r, "Plus T, P et B'g. Rien d'autre : cf. src/apps/bprime.cpp:304-330. Sans -char/-char-elem, "
                "bprime utilise un char de carbone pur intégré ; les deux options vont toujours ensemble.")
r += 1

r = section(ws, r, "§4 — LES QUATRE ROUTES DE CALCUL DU GAZ DE PYROLYSE")
rro = r + 1
r = headrow(ws, r, ["Route", "Matériau", "Donnée primaire", "Passage à l'élémentaire", "Point critique", "", ""])
for route, mat, prim, pas, crit in [
        ("A — sommation atomique", "TACOT, CPh70",
         "spéciation moléculaire mesurée (7 espèces)", "n_E = Σ_j ν_E,j · x_j",
         "la renormalisation : somme réelle des colonnes (100.30), pas les totaux imprimés (100.0)"),
        ("B — conversion directe", "ZURAM",
         "composition élémentaire MASSIQUE mesurée", "x_E = (y_E/M_E) / Σ (y_j/M_j)",
         "traçabilité limitée à 3 décimales ; la source primaire VKI n'est pas publique"),
        ("C — fermeture élémentaire", "SC-1008",
         "motif chimique idéalisé + rendement en char", "n_E(gaz) = n_E(résine) − n_E(char)",
         "le rendement en char entre DANS la chaîne : ±0.05 sur Y déplace le carbone de ~12 %"),
        ("D — fermeture sur DEUX constituants", "Liège/phénolique",
         "analyses des deux constituants + rendements", "n_E(gaz) = Σ_i [n_E(i) − n_E(char,i)]",
         "le renfort pyrolyse : le rapport renfort/résine entre dans la table B'"),
        ("E — stœchiométrie directe", "Carbone, Silice",
         "la formule du solide", "aucun gaz de pyrolyse (B'g = 0)",
         "le char EST le matériau ; pour la silice, -char-elem devient Si")]:
    r = row(ws, r, [route, mat, prim, pas, crit, "", ""],
            fonts=[F_LABB, F_LAB, F_LAB, F_MONO, F_NOTE, None, None])
    ws.row_dimensions[r-1].height = 28
    for c in "BCDEF":
        ws[f"{c}{r-1}"].alignment = Alignment(vertical="top", wrap_text=True)
r += 1

r = section(ws, r, "§5 — CE QUI N'ENTRE PAS DANS LE XML")
rno = r + 1
r = headrow(ws, r, ["Donnée", "Dans le XML ?", "Où elle sert vraiment", "", "", "", ""])
for d, inx, ou in [
        ("masse volumique ρ_v, ρ_c", "non", "solveur de réponse matériau ; récession ṡ = B'c·ṁe/ρ_c"),
        ("porosité", "non", "conduction, perméabilité"),
        ("rapport fibres/résine", "non *", "fixe la QUANTITÉ de gaz (donc B'g), pas sa composition"),
        ("rendement en char", "non *", "idem — mais il entre dans le CALCUL du gaz pour les routes C et D"),
        ("cinétique de pyrolyse", "non", "fixe B'g au cours du temps")]:
    r = row(ws, r, [d, inx, ou, "", "", "", ""], fonts=[F_LAB, F_LABB, F_NOTE, None, None, None, None])
r = note(ws, r, "* sauf si les constituants donnent des chars de compositions élémentaires DIFFÉRENTES "
                "(fibres de silice + résine phénolique), ou si le renfort pyrolyse (liège) : le rapport "
                "entre alors dans Y_c et/ou Y_g, et la table change. "
                "Le bilan de masse de surface est écrit sur des fractions massiques élémentaires, donc "
                "normalisées : les grandeurs VOLUMIQUES n'y ont structurellement aucune place.")
r += 1

r = section(ws, r, "§6 — RECETTE POUR UN NOUVEAU MATÉRIAU")
for n, txt in enumerate([
        "Atmosphère → composition -bl (air, Mars, Titan…).",
        "Résine → composition du gaz de pyrolyse par l'une des quatre routes du §4 → composition -py.",
        "Char → C:1.0 si tous les constituants carbonisent vers le carbone ; sinon pondérer par les masses de char de chaque constituant.",
        "Espèces → tous les produits attendus des éléments présents, PLUS la ou les phases condensées.",
        "Vérifier : checkmix (noms, phases), puis un run court.",
        "Contrôler la physique : à 300 K et B'g = 0 dans l'air, un char carboné doit donner B'c ≈ 0.0874 (limite C + O2 → CO2, indépendante de la pression). Une valeur ≈ 200 signale l'absence de C(gr).",
], 1):
    ws[f"A{r}"] = n
    ws[f"A{r}"].font = F_STEP
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = txt
    ws[f"B{r}"].font = F_LAB
    ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    r += 1
r += 1
r = para(ws, r, "Documents de référence du dépôt : tacot_bprime/mise_en_donnees_xml.md (mécanique du "
                "fichier XML et règles du parseur) · resine_tacot.md · zuram_bprime/resine_zuram.md · "
                "sc1008_bprime/resine_sc1008.md · cork_bprime/mise_en_donnees_cork.md · "
                "tacot_bprime/cph70_vs_tacot.md · carbon_bprime/bprime_carbon_physique.md.", 30)

# ---------------------------------------------------------------------------
wb.save(OUT)
print("écrit :", OUT)
